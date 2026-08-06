# -*- coding: utf-8 -*-
"""
红利数据更新工具（交互式 + 命令行）
场景：
  1. 日常更新   —— 行情增量（指数/ETF/推荐20/其他成份/自选）→ 前端包 → Excel → 变更摘要，约5分钟
  2. 深度更新   —— 日常 + 成分重下载→汇总表→分红检测→check-fin→回测(可选)，季度末推荐
  3. 单项操作   —— 只更新某一部分（子菜单）
  4. 维护工具   —— 回测重跑/失败重试/缓存清理/重导Excel
快捷命令：
  python update.py daily|full|idx|etf|rec|pool|watch|web|comp|summary|fin|bt|retry|excel|status [--yes]
用法: python update.py [命令]
"""
import sys, io, os, time, json
from datetime import datetime, date

BASE = os.path.dirname(os.path.abspath(__file__))
SCRIPTS = os.path.join(BASE, "scripts")
sys.path.insert(0, SCRIPTS)

import _fetch_history as fh
import _fetch_stock_data as fsd

# import后包装stdout（fh的包装对象仍被其模块引用，底层buffer不会被关闭）
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

INDICES = fh.INDICES
ETFS = fh.ETFS

CHANGE_LOG = os.path.join(BASE, "cache", "_change_log.json")   # 变更摘要状态（自动创建）
LAST_FIN = os.path.join(BASE, "cache", "_last_check_fin.json")  # check-fin 时间戳（自动创建）
WEB_DATA = os.path.join(BASE, "web", "data")

# ────────────────────────────── 通用小工具 ──────────────────────────────

def load_json(p, default=None):
    try:
        with open(p, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return default

def save_json(p, obj):
    tmp = p + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=1)
    os.replace(tmp, p)

def ask(msg, default=True):
    tip = "(y/n) " if default else "(y/N) "
    while True:
        r = input(f"⚠️  {msg}{tip}").strip().lower()
        if r in ("",):
            return default
        if r in ("y", "yes"):
            return True
        if r in ("n", "no"):
            return False
        print("  请输入 y 或 n")

def safe_export(desc, fn):
    """Excel 导出统一包裹：文件被占用时不中断流程，返回是否成功"""
    try:
        fn()
        return True
    except PermissionError:
        print(f"⚠️  {desc} 导出失败：文件被占用（请关闭 Excel 后重试该项）")
        return False

# ────────────────────────────── 变更摘要（_change_log.json） ──────────────────────────────

def _record_changes(key, cur):
    """采集变更日志：旧值挪到 {key}_prev，新值写入 {key}"""
    log = load_json(CHANGE_LOG) or {}
    log[f"{key}_prev"] = log.get(key)
    log[key] = cur
    log["time"] = time.strftime("%Y-%m-%d %H:%M")
    save_json(CHANGE_LOG, log)

def collect_indices(idx):
    """idx: _gen_components.build() 返回 {code: {name, stocks, date_cons, ...}}"""
    cur = {c: {"name": i.get("name", ""), "n": len(i.get("stocks") or {}),
               "sample": i.get("date_cons", "")} for c, i in idx.items()}
    _record_changes("indices", cur)

def collect_pool():
    n = len(load_json(os.path.join(BASE, "cache", "_成分股汇总.json")) or [])
    _record_changes("pool", n)

def collect_etfs():
    m = load_json(os.path.join(WEB_DATA, "manifest.json")) or {}
    cur = {e["code"]: {"name": e.get("name", "")} for e in m.get("etfs", [])}
    _record_changes("etfs", cur)

def print_changes():
    """本次 vs 上次 变更摘要（供人工核对《红利介绍.md》）"""
    log = load_json(CHANGE_LOG) or {}
    print("\n═══ 本次变更摘要 ═══")
    changes = []
    if "indices" in log and "indices_prev" in log:
        cur, prev = log["indices"], log["indices_prev"]
        for c, x in cur.items():
            p = prev.get(c)
            if not p:
                changes.append(f"新增指数 {c} {x['name']}（{x['n']}只）")
            elif p["n"] != x["n"]:
                changes.append(f"{c} {x['name']} 成分 {p['n']}→{x['n']}（{x['n']-p['n']:+d}），样本 {p['sample']}→{x['sample']}")
            elif p["sample"] != x["sample"]:
                changes.append(f"{c} {x['name']} 样本 {p['sample']}→{x['sample']}（成分数不变）")
        for c, p in prev.items():
            if c not in cur:
                changes.append(f"移除指数 {c} {p['name']}")
    if "etfs" in log and "etfs_prev" in log:
        cur, prev = log["etfs"], log["etfs_prev"]
        for c, x in cur.items():
            if c not in prev:
                changes.append(f"新增ETF {c} {x['name']}")
        for c, p in prev.items():
            if c not in cur:
                changes.append(f"移除ETF {c} {p['name']}")
    if "pool" in log and "pool_prev" in log and log["pool"] != log["pool_prev"]:
        changes.append(f"成分股汇总 {log['pool_prev']}→{log['pool']}（{log['pool']-log['pool_prev']:+d}）")
    if not changes:
        print("  无关键变化（成分/ETF/池规模均未变）")
    else:
        for c in changes:
            print(f"  · {c}")
    print("  ⚠️ 快照数字如有变化，请对照《红利介绍.md》同步")

# ────────────────────────────── 数据状态扫描（只读） ──────────────────────────────

def _d(s):
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None

def _failed_map():
    m = {}
    for p in ("cache/_pool_failed.json", "cache/_watchlist_failed.json"):
        for code, info in (load_json(os.path.join(BASE, p)) or {}).items():
            m[code] = info.get("err", "")
    return m

def scan_status():
    """只读扫描各类数据新鲜度（不写任何文件），输出状态面板"""
    m = load_json(os.path.join(WEB_DATA, "manifest.json")) or {}
    stocks = m.get("stocks", [])
    def pick(pred):
        s = next((s for s in stocks if pred(s) and s.get("ready")), None)
        return s and s.get("last")
    st = {
        "指数":      next((i.get("last") for i in m.get("indices", []) if i.get("last")), None),
        "ETF":       next((e.get("last") for e in m.get("etfs", []) if e.get("last")), None),
        "推荐20":    pick(lambda s: s.get("rec")),
        "其他成份":  pick(lambda s: not s.get("rec") and not s.get("watch")),
        "自选":      pick(lambda s: s.get("watch")),
    }
    cc = load_json(os.path.join(BASE, "cache", "成分_980092.json"))
    st["成分样本"] = cc and cc.get("sample_date")
    dates = [_d(v) for v in st.values() if v]
    base = max(dates) if dates else None
    print("\n═══ 数据状态 ═══")
    if base:
        for k, v in st.items():
            d = _d(v)
            if not d:
                print(f"  {k}: 无数据")
            elif d == base:
                print(f"  {k}: {v} ✓")
            else:
                print(f"  {k}: {v} ⚠️ 滞后 {(base-d).days} 天")
    else:
        for k in st:
            print(f"  {k}: 无数据")
    # 回测过期判定：backtest.date < analysis.date
    bt = load_json(os.path.join(WEB_DATA, "backtest.json"))
    an = load_json(os.path.join(WEB_DATA, "analysis.json"))
    if bt and an and bt.get("date") and an.get("date") and bt["date"] < an["date"]:
        print(f"  回测: {bt['date']} ⚠️ 已过期（维护→1 重跑，约1分钟）")
    # 季度检测
    lf = load_json(LAST_FIN)
    if lf and lf.get("time"):
        try:
            days = (date.today() - datetime.strptime(lf["time"][:10], "%Y-%m-%d").date()).days
        except Exception:
            days = None
        if days is not None:
            warn = "" if days <= 90 else " ⚠️ 建议 ≤90 天执行"
            print(f"  季度检测: 距上次 {days} 天{warn}")
    else:
        print("  季度检测: 从未执行（深度更新或维护可执行）")
    # 失败清单：92 开头=北交所预期失败，其余可重试
    fm = _failed_map()
    if fm:
        exp = [c for c in fm if c.startswith(("92", "83", "87", "88"))]
        rty = [c for c in fm if c not in exp]
        if rty:
            print(f"  失败清单: {len(fm)} 只（{len(rty)} 只可重试，维护→2 重试）")
        else:
            print(f"  失败清单: {len(fm)} 只（北交所无数据源，预期内）")
    print()

# ────────────────────────────── 各更新函数（原逻辑） ──────────────────────────────

def update_indices(incremental=True):
    print("\n═══ 指数历史行情更新 ═══")
    for code, name, src, tcode in INDICES:
        def fetcher(last_date, src=src, code=code, tcode=tcode):
            if src == "tencent":
                return fh.fetch_tencent_kline(tcode, code, start=last_date)
            if src == "csindex":
                start = last_date.replace("-", "") if last_date else None
                return fh.fetch_csindex_perf(code, start=start)
            return fh.fetch_cnindex_kline(code, start=last_date)
        try:
            fh.update_incremental("指数", code, name, fetcher)
        except Exception as e:
            print(f"  ❌ [{code} {name}] 失败: {repr(e)[:80]}")
    print("\n✅ 指数历史更新完成")

def update_etfs():
    print("\n═══ ETF历史行情更新（场内K线+净值）═══")
    for code, name, tcode in ETFS:
        def fetcher(last_date, code=code, tcode=tcode):
            kline = fh.fetch_tencent_kline(tcode, code, start=last_date)
            nav = fh.fetch_sina_nav(code, start=last_date)
            kmap = {r["date"]: r for r in kline}
            nmap = {r["date"]: r for r in nav}
            rows = []
            for d in sorted(set(kmap) | set(nmap)):
                r = {"date": d}
                if d in kmap:
                    r.update({k: v for k, v in kmap[d].items() if k != "date"})
                if d in nmap:
                    r["nav"] = nmap[d]["nav"]
                    r["acc_nav"] = nmap[d]["acc_nav"]
                rows.append(r)
            return rows
        try:
            fh.update_incremental("ETF", code, name, fetcher)
        except Exception as e:
            print(f"  ❌ [{code} {name}] 失败: {repr(e)[:80]}")
    print("\n✅ ETF历史更新完成")

def update_stocks():
    print("\n═══ 股票历史行情更新（20只，不复权，2004-01-01起）═══")
    fsd.update_dividends()   # 分红缓存缺失才拉（20只约1分钟），删除自愈
    fsd.update_financials()  # 财报缓存缺失才拉，删除自愈
    for code, name, tcode in fsd.STOCKS:
        try:
            fh.update_incremental("股票", code, name, fsd.make_fetcher(tcode, code))
        except Exception as e:
            print(f"  ❌ [{code} {name}] 失败: {repr(e)[:80]}")
    safe_export("推荐股指标Excel", fsd.export_excel)
    print("\n✅ 股票历史更新完成")

def export_excel():
    import pandas as pd
    from datetime import datetime as _dt
    from openpyxl import load_workbook
    from _fetch_history import load_cache

    def post_process_file(path):
        """导出后处理：日期列显示为年月日（保留日期类型）+ 美化（列宽/冻结首行/筛选/居中）"""
        wb = load_workbook(path)
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
                for cell in row:
                    if isinstance(cell.value, _dt):
                        cell.number_format = "yyyy-mm-dd"
            fh.beautify_sheet(ws)
        wb.save(path)
    # 展示口径：价格(点/元)、成交量(万手)、成交额(亿元)；单位净值/累计净值不加单位
    # 原始单位：腾讯 volume=手/amount=元(估算)；中证官网 tradingVol=股/tradingValue=亿元；国证 volume=万手/amount=亿元
    IDX_COL_CN = {
        "open": "开盘(点)", "close": "收盘(点)", "high": "最高(点)", "low": "最低(点)",
        "volume": "成交量(万手)", "amount": "成交额(亿元)",
        "chg30": "30日涨跌(%)", "chg60": "60日涨跌(%)", "chg90": "90日涨跌(%)",
    }
    IDX_DIV = {"tencent": (1e4, 1e8), "csindex": (1e6, 1), "cnindex": (1, 1)}  # (成交量除数, 成交额除数)
    ETF_COL_CN = {
        "open": "开盘(元)", "close": "收盘(元)", "high": "最高(元)", "low": "最低(元)",
        "volume": "成交量(万手)", "amount": "成交额(亿元)",
        "nav": "单位净值", "acc_nav": "累计净值",
        "chg30": "30日涨跌(%)", "chg60": "60日涨跌(%)", "chg90": "90日涨跌(%)",
    }
    print("\n═══ 重新生成 Excel ═══")
    # 指数
    idx_rows = {}
    for code, name, src, tcode in INDICES:
        c = load_cache("指数", code)
        if c:
            fh.fill_chg_n(c["rows"])   # 30/60/90 交易日涨跌幅（交易日口径）
            fh.save_cache("指数", code, c)
            idx_rows[code] = {"name": c.get("name", name), "source": src, "rows": c["rows"]}
    try:
        with pd.ExcelWriter(os.path.join(BASE, "excel", "指数历史.xlsx"), engine="openpyxl") as w:
            for code, info in idx_rows.items():
                rows = info["rows"]
                if not rows:
                    continue
                vdiv, adiv = IDX_DIV[info["source"]]
                df = pd.DataFrame(rows)
                df["date"] = pd.to_datetime(df["date"])
                df = df.sort_values("date").set_index("date")
                df = df.rename(columns=IDX_COL_CN)
                df = df[list(IDX_COL_CN.values())]
                # 全 None 列会被推断为 object，强制数值化（腾讯源指数无成交额→NaN）
                df["成交量(万手)"] = pd.to_numeric(df["成交量(万手)"], errors="coerce")
                df["成交额(亿元)"] = pd.to_numeric(df["成交额(亿元)"], errors="coerce")
                df["成交量(万手)"] = (df["成交量(万手)"] / vdiv).round(2)
                df["成交额(亿元)"] = (df["成交额(亿元)"] / adiv).round(2)
                df.index.name = "日期"
                df.to_excel(w, sheet_name=f"{code} {info['name'][:10]}"[:31])
        post_process_file(os.path.join(BASE, "excel", "指数历史.xlsx"))
    except PermissionError:
        print("⚠️  excel/指数历史.xlsx 被占用（可能已在Excel中打开），请关闭后重新执行导出")
    # ETF（先估算成交额并写回缓存）
    etf_rows = {}
    for code, name, tcode in ETFS:
        c = load_cache("ETF", code)
        if c:
            fh.fill_etf_amount(c["rows"])
            fh.fill_chg_n(c["rows"])   # 30/60/90 交易日涨跌幅（交易日口径）
            fh.save_cache("ETF", code, c)
            etf_rows[code] = {"name": c.get("name", name), "rows": c["rows"]}
    try:
        with pd.ExcelWriter(os.path.join(BASE, "excel", "ETF历史.xlsx"), engine="openpyxl") as w:
            for code, info in etf_rows.items():
                rows = info["rows"]
                if not rows:
                    continue
                df = pd.DataFrame(rows)
                df["date"] = pd.to_datetime(df["date"])
                df = df.sort_values("date").set_index("date")
                df = df.rename(columns=ETF_COL_CN)
                df["成交量(万手)"] = pd.to_numeric(df["成交量(万手)"], errors="coerce")
                df["成交额(亿元)"] = pd.to_numeric(df["成交额(亿元)"], errors="coerce")
                df["成交量(万手)"] = (df["成交量(万手)"] / 1e4).round(2)
                df["成交额(亿元)"] = (df["成交额(亿元)"] / 1e8).round(2)
                if "单位净值" in df.columns:
                    df = df[["开盘(元)", "收盘(元)", "最高(元)", "最低(元)", "成交量(万手)", "成交额(亿元)",
                             "单位净值", "累计净值", "30日涨跌(%)", "60日涨跌(%)", "90日涨跌(%)"]]
                df.index.name = "日期"
                df.to_excel(w, sheet_name=f"{code} {info['name'][:10]}"[:31])
        post_process_file(os.path.join(BASE, "excel", "ETF历史.xlsx"))
    except PermissionError:
        print("⚠️  excel/ETF历史.xlsx 被占用（可能已在Excel中打开），请关闭后重新执行导出")
    else:
        print("✅ excel/ETF历史.xlsx 已更新")

def update_components():
    print("\n═══ 指数成分股更新 ═══")
    import _gen_components  # scripts/ 已在 sys.path
    idx = _gen_components.build()
    # 摘要
    print("\n更新摘要：")
    for code in ["000922", "000015", "000821", "000825", "H30269", "930955", "932315", "931468", "H30270", "980092", "000151"]:
        info = idx.get(code, {})
        stocks = info.get("stocks") or {}
        fb = " ⚠️旧数据回退" if info.get("fallback") else ""
        print(f"  {code} {info.get('name','')}: {len(stocks)}只, 样本截止 {info.get('date_cons','')}{fb}")
    collect_indices(idx)   # 变更摘要采集
    print("\n⚠️ 请手动核对《红利介绍.md》中相关描述/快照数字是否需要同步（脚本不自动改该文件）。")

def update_summary(force=False):
    print("\n═══ 成分股汇总表更新（解析md→行业/行情/股息率→汇总Excel）═══")
    import _update_summary
    _update_summary.run(force=force)
    collect_pool()   # 变更摘要采集

def update_web():
    """前端数据包：web/data/ 四件套 + 买卖区间分析（S1反推+S3/S4打分） + 国证指数成分"""
    print("\n═══ 前端数据包更新（web/data/ + 区间分析 + 国证成分）═══")
    import _gen_web_data as gwd
    gwd.build_manifest()
    gwd.build_stock_indicators()
    gwd.build_components()
    gwd.build_summary()
    import _fetch_etf_holdings as feh
    feh.main()   # ETF季报持仓（980092 股息率估算、159201 持仓成分）
    import _gen_analysis as ga
    ga.main()   # S1 股息率反推（重建 analysis_dy.json）→ S3/S4 因子打分与点位锚（analysis.json）
    import _fetch_cnindex_components as fcc
    fcc.main()
    collect_etfs()   # 变更摘要采集
    print("\n⚠️ 提示：回测报告（web/data/backtest.json + docs/回测报告.md）为研究产物，")
    print("   需手动运行 python scripts/_backtest_analysis.py 更新（约1分钟，数据更新后建议重跑）。")
    print("⚠️ 请人工核对《红利介绍.md》中相关描述/快照数字是否需要同步（脚本不自动改该文件）。")
    print("✅ 前端数据包更新完成（刷新浏览器即可看到新数据）")

def update_pool(rerun_web=True):
    """其他成份股（汇总表 − 推荐 20）：K线增量 + 缺失补齐 → 重算指标与区间分析"""
    print("\n═══ 其他成份股更新（首次全量约 20-30 分钟，之后增量约 3 分钟）═══")
    import _fetch_pool_data as fpd
    fpd.main()
    if rerun_web:
        print("\n── 重算前端数据包（指标 + 区间分析）──")
        import _gen_web_data as gwd
        gwd.build_stock_indicators()   # 先指标（manifest 的 last_dy 读指标文件末行）
        gwd.build_manifest()
        import _gen_analysis as ga
        ga.main()
        print("\n⚠️  季度末请执行：python scripts/_fetch_pool_data.py --check-fin（分红/财报/股本检测，约15分钟）")
        print("✅ 其他成份股更新完成（刷新浏览器即可）")


def update_watchlist(rerun_web=True):
    """自选股（读自选股清单.xlsx）：K线/分红/财报/股本增量 → 重算前端包"""
    print("\n═══ 自选股更新（自选股清单.xlsx）═══")
    import _fetch_watchlist as fw
    fw.main()
    if rerun_web:
        print("\n── 重算前端数据包（清单 + 指标 + 区间分析）──")
        import _gen_web_data as gwd
        gwd.build_stock_indicators()   # 先指标（manifest 的 last_dy 读指标文件末行）
        gwd.build_manifest()
        import _gen_analysis as ga
        ga.main()
        print("✅ 自选股更新完成（刷新浏览器即可）")


def update_fin_refresh():
    print("\n═══ 分红/财报更新检测（季度末/定期执行）═══")
    fsd.check_financials()
    safe_export("推荐股指标Excel", fsd.export_excel)   # 指标列（股息率/PE等）依赖财报，检测后重导出
    print("✅ 分红/财报检测完成")

# ────────────────────────────── 维护工具 ──────────────────────────────

def run_backtest():
    print("\n═══ 重跑回测（精选池，约1分钟）═══")
    import _backtest_analysis as ba
    ba.main()
    print("✅ 回测完成（web/data/backtest.json + docs/回测报告.md）")

def retry_failed():
    print("\n═══ 失败清单重试（跳过北交所预期失败项）═══")
    import _fetch_pool_data as fpd
    fpd.main(retry_failed=True)
    import _fetch_watchlist as fw
    fw.main(retry_failed=True)
    print("\n── 重算前端数据包 ──")
    import _gen_web_data as gwd
    gwd.build_stock_indicators()
    gwd.build_manifest()
    import _gen_analysis as ga
    ga.main()
    print("✅ 重试完成（刷新浏览器即可）")

def clear_cache():
    """删除指定标的缓存 → 下次更新全量重拉（删除自愈）"""
    print("\n═══ 缓存清理（删除后对应更新将全量重拉）═══")
    print("  可输入：指数/ETF 代码（如 000922 512890）、股票代码（如 601398），空格分隔多个")
    print("  股票会连带删除其 分红/财报/股本 缓存；输入 0 返回")
    codes = input("请输入代码: ").strip().split()
    if not codes or codes == ["0"]:
        return
    hit = []
    for c in codes:
        c = c.strip().upper()
        for typ, pat in (("指数", f"指数_{c}.json"), ("ETF", f"ETF_{c}.json")):
            p = os.path.join(BASE, "cache", pat)
            if os.path.exists(p):
                os.remove(p)
                hit.append(f"cache/{pat}")
        for pat in (f"股票_{c}.json", f"分红_{c}.json", f"财报_{c}.json", f"股本_{c}.json"):
            p = os.path.join(BASE, "cache", pat)
            if os.path.exists(p):
                os.remove(p)
                hit.append(f"cache/{pat}")
    if hit:
        print("  已删除:")
        for h in hit:
            print(f"    {h}")
        print("  → 请运行对应更新（单项/日常）以全量重拉")
    else:
        print("  未找到匹配的缓存文件（检查代码是否正确）")

# ────────────────────────────── 编排（日常/深度） ──────────────────────────────

def daily_update(auto=False):
    """日常：行情增量（指数/ETF/推荐20/其他/自选）→ Excel → 前端包 → 变更摘要"""
    if not auto:
        if not ask("日常更新（行情增量→Excel→前端包，约5分钟）是否执行？"):
            return
    t0 = time.time()
    print("\n" + "═" * 50 + "\n  日常更新（行情增量 → Excel → 前端包）\n" + "═" * 50)
    update_indices()
    update_etfs()
    update_stocks()
    update_pool(rerun_web=False)
    update_watchlist(rerun_web=False)
    export_excel()
    update_web()
    print_changes()
    print(f"\n✅ 日常更新完成，用时 {time.time()-t0:.0f} 秒（刷新浏览器即可）")

def deep_update(auto=False):
    """深度：日常行情 + 成分重下→汇总表→分红检测→check-fin(询问)→回测(询问)"""
    if not auto:
        if not ask("深度更新（日常 + 成分重下载/汇总表/分红检测/check-fin，约20分钟）是否执行？"):
            return
    t0 = time.time()
    print("\n" + "═" * 50 + "\n  深度更新（季度链路：3→4→8→check-fin→9→10→web）\n" + "═" * 50)
    update_indices()
    update_etfs()
    update_stocks()
    update_components()     # 3 重下载样本 → 重写成分md
    update_summary()        # 4 解析md → 汇总表（新成分自动进池）
    update_fin_refresh()    # 8 推荐20分红/财报检测
    if ask("check-fin（其他成份股分红/财报/股本季度检测，约15分钟）是否执行？", default=False):
        import _fetch_pool_data as fpd
        fpd.main(check_fin=True)
        import _fetch_watchlist as fw
        fw.main(check_fin=True)
        save_json(LAST_FIN, {"time": time.strftime("%Y-%m-%d %H:%M")})
    update_pool(rerun_web=False)      # 9 其他成份（含新成分K线）
    update_watchlist(rerun_web=False)  # 10 自选
    export_excel()
    update_web()
    if ask("是否重跑回测（约1分钟）？", default=False):
        run_backtest()
    print_changes()
    print(f"\n✅ 深度更新完成，用时 {time.time()-t0:.0f} 秒")

# ────────────────────────────── 菜单 ──────────────────────────────

def single_menu():
    while True:
        print("""
── 单项操作 ──
  1. 指数历史    2. ETF历史    3. 推荐20只
  4. 其他成份股  5. 自选股     6. 前端数据包
  7. 成分重下载  8. 汇总表     9. 分红/财报检测
  0. 返回
""")
        ch = input("请输入编号: ").strip()
        if ch == "0":
            return
        elif ch == "1":
            update_indices(); export_excel()
        elif ch == "2":
            update_etfs(); export_excel()
        elif ch == "3":
            update_stocks()
        elif ch == "4":
            update_pool()
        elif ch == "5":
            update_watchlist()
        elif ch == "6":
            update_web()
        elif ch == "7":
            update_components()
        elif ch == "8":
            update_summary()
        elif ch == "9":
            update_fin_refresh()
        else:
            print("无效输入")

def tools_menu():
    while True:
        print("""
── 维护工具 ──
  1. 重跑回测（backtest.json + 回测报告.md，约1分钟）
  2. 失败重试（跳过北交所预期失败项）
  3. 缓存清理（删除后全量重拉）
  4. 重导 Excel（指数/ETF/推荐股）
  5. 季度检测 check-fin（约15分钟）
  0. 返回
""")
        ch = input("请输入编号: ").strip()
        if ch == "0":
            return
        elif ch == "1":
            run_backtest()
        elif ch == "2":
            retry_failed()
        elif ch == "3":
            clear_cache()
        elif ch == "4":
            export_excel()
            safe_export("推荐股指标Excel", fsd.export_excel)
        elif ch == "5":
            import _fetch_pool_data as fpd
            fpd.main(check_fin=True)
            import _fetch_watchlist as fw
            fw.main(check_fin=True)
            save_json(LAST_FIN, {"time": time.strftime("%Y-%m-%d %H:%M")})
            print("✅ check-fin 完成（建议随后运行 单项→4 重算指标）")
        else:
            print("无效输入")

def main_menu():
    print("""
请选择：
  1. 日常更新   —— 行情增量（指数/ETF/推荐20/其他/自选）→ 前端包 → Excel → 变更摘要，约5分钟
  2. 深度更新   —— 日常 + 成分重下载→汇总表→分红检测→check-fin→回测(可选)，季度末推荐
  3. 单项操作   —— 只更新某一部分（子菜单）
  4. 维护工具   —— 回测重跑/失败重试/缓存清理/重导Excel
  0. 退出
""")
    ch = input("请输入编号: ").strip()
    if ch == "1":
        daily_update(auto=True)
    elif ch == "2":
        deep_update(auto=True)
    elif ch == "3":
        single_menu()
    elif ch == "4":
        tools_menu()
    elif ch == "0":
        print("已退出")
        return False
    else:
        print("无效输入")
    return True

# ────────────────────────────── 命令行直跑 ──────────────────────────────

CMDS = {
    "daily":   lambda: daily_update(auto="--yes" in sys.argv),
    "full":    lambda: deep_update(auto="--yes" in sys.argv),
    "idx":     lambda: (update_indices(), export_excel()),
    "etf":     lambda: (update_etfs(), export_excel()),
    "rec":     update_stocks,
    "pool":    update_pool,
    "watch":   update_watchlist,
    "web":     update_web,
    "comp":    update_components,
    "summary": update_summary,
    "fin":     update_fin_refresh,
    "bt":      run_backtest,
    "retry":   retry_failed,
    "excel":   export_excel,
    "status":  scan_status,
}

def run_cmd(args):
    name = next((a for a in args if not a.startswith("--")), "")
    if not name:
        print(f"用法: python update.py {'|'.join(CMDS)} [--yes]")
        return
    if name not in CMDS:
        print(f"未知命令: {name}（可用: {' '.join(CMDS)}）")
        return
    CMDS[name]()

def main():
    args = sys.argv[1:]
    if args:
        run_cmd(args)
        return
    print("═" * 50)
    print("  红利数据更新工具")
    print("═" * 50)
    print(f"快捷: python update.py {' | '.join(CMDS)}")
    while True:
        scan_status()
        if not main_menu():
            break

if __name__ == "__main__":
    main()
