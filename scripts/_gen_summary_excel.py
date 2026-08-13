# -*- coding: utf-8 -*-
"""生成 红利成分股汇总.xlsx：综合 md解析权重 + 东财行业 + 腾讯估值 + 分红股息率 全字段，中文表头"""
import json, sys, os, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

def main():
    for p in ("cache/_成分股汇总.json", "cache/_成分股汇总表.json"):
        if not os.path.exists(p):
            sys.exit(f"❌ 缺少 {p}，请先运行 update.py 选项4（解析成分股md→行业/行情/股息率→缓存→Excel）")

    stock = json.load(open("cache/_成分股汇总.json", encoding="utf-8"))
    table = json.load(open("cache/_成分股汇总表.json", encoding="utf-8"))
    # 股息率/分红记录已全量写入 _成分股汇总.json（289只），不再依赖65只候选文件
    cand_path = "cache/_候选股息率.json"
    dy_map = {r["code"]: r for r in json.load(open(cand_path, encoding="utf-8"))} if os.path.exists(cand_path) else {}
    # 以全量缓存为准，候选文件仅作交叉验证
    for t in table:
        s = stock[t["code"]]
        if t["div_yield"] is None and s.get("div_yield_calc") is not None:
            t["div_yield"] = s["div_yield_calc"]
        if not t.get("div_rec") and s.get("div_rec"):
            t["div_rec"] = s["div_rec"]

    # 20只最终推荐（动态：读评分产物 cache/_推荐20.json 的 list 段=TOP20；缺失回退硬编码）
    _FALLBACK_FINAL20 = {"600036","601838","601088","601225","600938","601857","600350","601006",
                         "600900","600795","000858","000895","000651","000333","000423","600566",
                         "600019","601668","600582","600757"}
    def _final20():
        try:
            r = json.load(open("cache/_推荐20.json", encoding="utf-8"))
            lst = r.get("list")
            if lst:
                return {x["code"] for x in lst}
        except Exception:
            pass
        print("  ⚠️ cache/_推荐20.json 缺失/损坏，最终推荐列回退硬编码清单")
        return set(_FALLBACK_FINAL20)
    FINAL20 = _final20()

    rows = []
    for t in table:
        c = t["code"]
        s = stock[c]
        d = dy_map.get(c, {})
        # 主要入选指数及权重（前8个，格式: 指数名 权重%）；权重未公开的标注
        idx_str = "；".join(
            f"{k} {v}%" if v is not None else f"{k}(权重未公开)"
            for k, v in t["idx"][:8]
        )
        # 分红记录（每10股口径）—— 优先全量缓存，回退候选文件
        rec = t.get("div_rec") or dy_map.get(c, {}).get("div_rec") or []
        rec_str = "；".join(f"{ex}派{bonus}元/10股" for ex, bonus in rec[:5]) if rec else "近12个月无派息记录"
        rows.append({
            "代码": c,
            "名称": s["name"],
            "一级行业": t["ind"],
            "细分行业": t["ind3"],
            "入选指数/ETF数": t["n"],
            "最大权重%": t["maxw"],
            "最新价": t["price"],
            "涨跌幅%": s.get("change_pct"),
            "PE(TTM)": t["pe"],
            "PB": t["pb"],
            "总市值(亿)": round(t["mcap"], 0),
            "近12个月股息率%": t.get("div_yield") if t.get("div_yield") is not None else (d.get("div_yield") if d else None),        "近5次分红记录": rec_str,
            "主要入选指数及权重": idx_str,
            "入选20只推荐": "是" if c in FINAL20 else "否",
        })

    # 排序：一级行业 → 入选数 → 最大权重
    rows.sort(key=lambda r: (r["一级行业"], -r["入选指数/ETF数"], -r["最大权重%"]))

    # ── 写 Excel ──
    wb = Workbook()
    ws = wb.active
    ws.title = "成分股汇总"

    headers = ["序号", "证券代码", "证券名称", "一级行业", "细分行业", "入选指数/ETF数", "最大权重(%)",
               "最新价(元)", "当日涨跌幅(%)", "PE(TTM)", "PB", "总市值(亿元)", "近12个月股息率(%)",
               "近5次分红记录", "主要入选指数及权重(%)", "是否入选20只推荐"]

    # 样式
    head_fill = PatternFill("solid", fgColor="C00000")
    head_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    rec_fill = PatternFill("solid", fgColor="FFF2CC")   # 推荐行浅黄
    alt_fill = PatternFill("solid", fgColor="F7F7F7")   # 隔行浅灰

    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = center
        cell.border = border

    for i, r in enumerate(rows, 1):
        vals = [i, r["代码"], r["名称"], r["一级行业"], r["细分行业"], r["入选指数/ETF数"], r["最大权重%"],
                r["最新价"], r["涨跌幅%"], r["PE(TTM)"], r["PB"], r["总市值(亿)"], r["近12个月股息率%"],
                r["近5次分红记录"], r["主要入选指数及权重"], r["入选20只推荐"]]
        is_rec = r["入选20只推荐"] == "是"
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=i + 1, column=j, value=v)
            cell.border = border
            if j in (2, 3):
                cell.alignment = left
            elif j in (14, 15):
                cell.alignment = left
            else:
                cell.alignment = center
            if is_rec:
                cell.fill = rec_fill
            elif i % 2 == 0:
                cell.fill = alt_fill
        # 数值格式
        for j in (8, 9, 10, 11, 13):
            cell = ws.cell(row=i + 1, column=j)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"
        cell = ws.cell(row=i + 1, column=12)
        if isinstance(cell.value, (int, float)):
            cell.number_format = "#,##0"

    # 列宽
    widths = [6, 10, 12, 10, 12, 12, 10, 10, 12, 9, 8, 12, 13, 46, 60, 12]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:P{len(rows) + 1}"

    # 封面说明 sheet
    ws2 = wb.create_sheet("说明")
    notes = [
        f"红利成分股汇总表（数据日期：{datetime.date.today()}）",
        "",
        "数据来源：",
        "1. 成分与权重：《红利指数与ETF成分股.md》解析（精选指数10只+ETF 11只，权重为2026-06-30/07-31静态快照）",
        "2. 行业：东财个股接口（申万行业分类），一级行业为归并后的申万大类",
        "3. 行情/估值：腾讯财经公开接口（2026-08-02收盘）",
        "4. 股息率：东财分红历史接口，近12个月每股派息合计÷最新价（全部289只均已计算；近12个月无派息者记 0.00%）",
        "5. 入选20只推荐：指《红利股票推荐20只.md》最终组合（每行业≤2只）",
        "",
        "列说明：",
        "· 入选指数/ETF数：该股票出现在精选池（10只指数+11只ETF，ETF与跟踪指数成分一致）中的个数，最大21",
        "· 最大权重(%)：该股票在所有入选指数/ETF中的最高权重",
        "· 近5次分红记录：格式为【除权日+每10股税前派息(元)】",
        "· 主要入选指数及权重：按权重降序的前8条，格式【指数名 权重%】；权重未公开（如中证800自由现金流/上证国有企业红利）标注【指数名(权重未公开)】",
        "· PE(TTM) 为负：该股近12个月处于亏损状态（如中集集团），为真实数据",
        "",
        "风险提示：以上内容仅为基于公开数据的客观信息梳理，不构成投资建议。",
    ]
    for i, line in enumerate(notes, 1):
        ws2.cell(row=i, column=1, value=line)
    ws2.column_dimensions["A"].width = 110

    out = "excel/红利成分股汇总.xlsx"
    wb.save(out)
    print(f"已生成 {out}：{len(rows)} 只股票 × {len(headers)} 列")
    from collections import Counter
    print("行业分布:", dict(Counter(r["一级行业"] for r in rows)))
    print("推荐20只覆盖:", sum(1 for r in rows if r["入选20只推荐"] == "是"))


if __name__ == "__main__":
    main()
