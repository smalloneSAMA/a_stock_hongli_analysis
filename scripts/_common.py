# -*- coding: utf-8 -*-
"""公共工具单一事实来源：行情批量 / 东财限流 / 原子读写 / Excel 导出（P1.5 补）。

新脚本必须从此 import，禁止本地复制（AGENTS.md 约束）。
用法: python scripts/_common.py   # 冒烟自测（不触网）
"""
import json, os, random, time, urllib.request

# ── 市场前缀路由（唯一正确版本；源：_fetch_watchlist.py:132-136）──
# 顺序有依赖勿调整：92(北交所)必须早于 9(沪B)；4/8(北交所旧号段)
def market_prefix(code):
    """股票/ETF 代码 → 腾讯行情带前缀全名（如 600036→sh600036、920599→bj920599）"""
    if code.startswith("92"):
        return "bj" + code
    if code.startswith(("5", "6", "9")):
        return "sh" + code
    if code.startswith(("4", "8")):
        return "bj" + code
    return "sz" + code

# ── 腾讯批量行情 ──
TENCENT_URL = "https://qt.gtimg.cn/q="
TENCENT_BATCH = 50      # AGENTS.md：50只/批防封（曾用 60 的副本已统一）
V_NAME, V_PRICE = 1, 3
V_PE_TTM, V_PB = 39, 46
V_MCAP_YI = 45          # 总市值，单位已是亿元，勿再除 1e8（困难总结 v[45] 事故）

def tencent_quotes(codes, timeout=10, retries=2):
    """批量实时行情 → {code: v字段list}（原始字段，调用处自行提取所需位）。

    口径与 _fetch_watchlist.py:139-164 一致：50只/批、批间 0.3s、gbk errors=replace、
    len(v)<53 或无行情行跳过；批次失败重试 retries 次，仍失败仅告警不中断（其余批次照常）。"""
    out = {}
    for i in range(0, len(codes), TENCENT_BATCH):
        batch = codes[i:i + TENCENT_BATCH]
        url = TENCENT_URL + ",".join(market_prefix(c) for c in batch)
        data = None
        for attempt in range(retries + 1):
            try:
                req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
                data = urllib.request.urlopen(req, timeout=timeout).read().decode("gbk", errors="replace")
                break
            except Exception as e:
                if attempt == retries:
                    print(f"  ❌ 腾讯行情批次 {i // TENCENT_BATCH + 1} 失败: {repr(e)[:70]}")
                else:
                    time.sleep(2 + attempt * 2)
        if data is None:
            time.sleep(0.3)
            continue
        for line in data.strip().split(";"):
            if '"' not in line:
                continue
            key = line.split("=")[0].split("_")[-1]
            v = line.split('"')[1].split("~")
            if len(v) < 53:
                continue
            out[key[2:]] = v
        time.sleep(0.3)
    return out


def parse_quote(v):
    """v 字段list → {name, price, pe, pb, mcap_yi}；缺失/0 → None（口径：_fetch_watchlist.py:153-160）"""
    def f(i):
        try:
            x = float(v[i] or 0)
        except (ValueError, IndexError):
            return None
        return round(x, 2) if x else None
    return {"name": v[V_NAME] if len(v) > V_NAME else "",
            "price": f(V_PRICE), "pe": f(V_PE_TTM), "pb": f(V_PB), "mcap_yi": f(V_MCAP_YI)}


# ── 东财接口限流请求（1s/请求防封，全局限流）──
EM_UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/126.0 Safari/537.36"
_last_em = [0.0]

def em_get(url, timeout=12):
    """东财接口 GET → str（utf-8）。全局限流 1s/请求 + 0.1~0.4s 抖动（源：_fetch_stock_data.py:91-100）"""
    wait = 1.0 - (time.time() - _last_em[0])
    if wait > 0:
        time.sleep(wait + random.uniform(0.1, 0.4))
    req = urllib.request.Request(url, headers={"User-Agent": EM_UA, "Referer": "https://quote.eastmoney.com/"})
    try:
        return urllib.request.urlopen(req, timeout=timeout).read().decode()
    finally:
        _last_em[0] = time.time()


# ── 原子读写（缓存保护三件套之一：原子写 tmp+replace）──
def atomic_dump(path, obj, indent=1, separators=None):
    """原子写 JSON：先写 path.tmp 再 os.replace，中断不损坏缓存"""
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=indent, separators=separators)
    os.replace(tmp, path)


def atomic_load(path, default=None):
    """读 JSON：不存在/解码失败返回 default（损坏自愈入口，P2 全面启用）"""
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except (OSError, ValueError):
        return default


# ── Excel 导出统一包裹（文件被占用不中断流程；源：update.py:60-67）──
def safe_export(desc, fn):
    """执行 fn() 导出 Excel；PermissionError（文件被 Excel 占用）时告警并返回 False 不中断"""
    try:
        fn()
        return True
    except PermissionError:
        print(f"⚠️  {desc} 导出失败：文件被占用（请关闭 Excel 后重试该项）")
        return False


# ── Excel 导出统一（万手/亿元 展示口径；P1.5 三合一，源：_fetch_history.beautify_sheet / update.py post_process）──
def beautify_sheet(ws):
    """单个sheet：列宽按内容显示宽度自适应、冻结首行、开启筛选、全表水平垂直居中，表头加粗浅蓝底。
    幂等，可重复执行"""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    import datetime as _dt
    if ws.max_row < 1 or ws.max_column < 1:
        return
    nrow, ncol = ws.max_row, ws.max_column

    def dispw(v):
        if isinstance(v, _dt.datetime):
            return 10   # 日期按 yyyy-mm-dd 显示宽度计
        s = str(v)
        return sum(2 if ord(ch) > 127 else 1 for ch in s)   # 中文/全角按2字符宽

    widths = [0] * ncol
    for row in ws.iter_rows(min_row=1, max_row=nrow, max_col=ncol):
        for cell in row:
            if cell.value is None:
                continue
            w = dispw(cell.value)
            i = cell.column - 1
            if w > widths[i]:
                widths[i] = w
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 8), 22)

    ws.freeze_panes = "A2"            # 冻结首行
    ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}{nrow}"   # 开启筛选

    center = Alignment(horizontal="center", vertical="center")
    hfont = Font(bold=True)
    hfill = PatternFill("solid", fgColor="D9E1F2")
    for row in ws.iter_rows(min_row=1, max_row=nrow, max_col=ncol):
        for cell in row:
            if cell.value is None:
                continue
            cell.alignment = center
            if cell.row == 1:
                cell.font = hfont
                cell.fill = hfill


def beautify_workbook(path):
    """导出后处理：日期列显示为年月日 + 逐sheet美化（列宽/冻结/筛选/居中）。
    幂等；导出后独立执行，避免 pandas 保存覆盖格式（源：update.py post_process_file）"""
    from datetime import datetime as _dt
    from openpyxl import load_workbook
    wb = load_workbook(path)
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
            for cell in row:
                if isinstance(cell.value, _dt):
                    cell.number_format = "yyyy-mm-dd"
        beautify_sheet(ws)
    wb.save(path)


def export_workbook(path, sheets, post=True):
    """统一 Excel 导出：sheets = {sheet名: DataFrame}（已含日期索引+中文列名+单位换算）。
    post=True 时导出后统一日期格式与美化。文件被占用抛 PermissionError，由 safe_export 包裹。"""
    import pandas as pd
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        for name, df in sheets.items():
            df.to_excel(w, sheet_name=name[:31])
    if post:
        beautify_workbook(path)


if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding="utf-8")   # Windows GBK 控制台兼容（不换对象，避免二次包装坑）
    fails = 0
    def chk(name, cond, detail=""):
        global fails
        if cond:
            print(f"  ✅ {name}")
        else:
            fails += 1
            print(f"  ❌ {name} {detail}")
    print("═══ _common.py 冒烟自测（不触网）═══")
    # ── market_prefix 用例表（困难总结 #11/#23：92 必须先于 9x）──
    chk("600036→sh600036", market_prefix("600036") == "sh600036")
    chk("000858→sz000858", market_prefix("000858") == "sz000858")
    chk("920599→bj920599（92先于9x）", market_prefix("920599") == "bj920599")
    chk("430047→bj430047", market_prefix("430047") == "bj430047")
    chk("830799→bj830799", market_prefix("830799") == "bj830799")
    chk("900901→sh900901（沪B）", market_prefix("900901") == "sh900901")
    chk("510880→sh510880（沪基金）", market_prefix("510880") == "sh510880")
    # ── parse_quote fixture（腾讯 v 字段样例，关键位对齐）──
    v = [""] * 53
    v[1], v[3], v[39], v[45], v[46] = "招商银行", "33.50", "6.50", "9785.30", "1.05"
    q = parse_quote(v)
    chk("name", q["name"] == "招商银行")
    chk("price", q["price"] == 33.5)
    chk("pe(TTM)", q["pe"] == 6.5)
    chk("mcap_yi 已是亿元勿再除", q["mcap_yi"] == 9785.3)
    chk("pb", q["pb"] == 1.05)
    q0 = parse_quote([""] * 53)
    chk("缺失/0→None", q0["pe"] is None and q0["mcap_yi"] is None and q0["price"] is None)
    # ── atomic 读写往返 ──
    import tempfile
    tp = os.path.join(tempfile.gettempdir(), "_common_smoke.json")
    atomic_dump(tp, {"a": 1, "b": [2, 3]})
    chk("atomic_dump→atomic_load 往返", atomic_load(tp) == {"a": 1, "b": [2, 3]})
    chk("atomic_load 缺失返 default", atomic_load(tp + ".none") is None)
    chk("atomic_load 损坏返 default", atomic_load(__file__) is None)  # py 源码非 JSON
    os.remove(tp)
    print("═══ 汇总：%s ═══" % ("PASS" if fails == 0 else "FAIL %d" % fails))
    raise SystemExit(1 if fails else 0)
