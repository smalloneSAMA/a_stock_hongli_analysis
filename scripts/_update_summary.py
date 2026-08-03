# -*- coding: utf-8 -*-
"""
成分股汇总数据更新模块（供 update.py 调用，也可单独运行）

链路：解析《红利指数与ETF成分股.md》 → 增量补行业(push2delay)/行情(腾讯)/分红股息率(datacenter)
      → 行业归并 → 生成 cache 汇总JSON → 生成 excel/红利成分股汇总.xlsx → 核对推荐20只

增量策略（避免长时间等待）：
  · 行业：仅拉 md 中新增的股票（已有缓存跳过）
  · 行情：腾讯批量全量刷新（不封IP，快）
  · 股息率/分红：默认仅补新增股票；--force 全量重算（约6分钟）
  · 连续8次失败自动中止（防IP被封后空等）
"""
import json, os, sys, re, time, random, datetime, urllib.request

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CACHE = os.path.join(BASE, "cache")
EXCEL = os.path.join(BASE, "excel")
MD = os.path.join(BASE, "红利指数与ETF成分股.md")
SUMMARY_JSON = os.path.join(CACHE, "_成分股汇总.json")
TABLE_JSON = os.path.join(CACHE, "_成分股汇总表.json")

def atomic_dump(path, obj):
    """原子写 JSON：先写 .tmp 再 os.replace，防止写一半中断损坏缓存"""
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False)
    os.replace(tmp, path)

# 20只推荐名单（《红利股票推荐20只.md》）
FINAL20 = ["600036", "601838", "601088", "601225", "600938", "601857", "600350", "601006",
           "600900", "600795", "000858", "000895", "000651", "000333", "000423", "600566",
           "600019", "601668", "600582", "600757"]

UA = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0",
      "Referer": "https://quote.eastmoney.com/"}
_last = [0.0]
FAIL_LIMIT = 8          # 连续失败上限

def em_get(url, timeout=8):
    wait = 1.0 - (time.time() - _last[0])
    if wait > 0:
        time.sleep(wait + random.uniform(0.1, 0.4))
    req = urllib.request.Request(url, headers=UA)
    try:
        return urllib.request.urlopen(req, timeout=timeout).read().decode()
    finally:
        _last[0] = time.time()

# ── 1. 解析成分股md → 基础缓存（name/w/n）────────────────────────────
def parse_components():
    if not os.path.exists(MD):
        print(f"  ❌ 找不到《红利指数与ETF成分股.md》（{MD}），请先运行 update.py 选项3 生成")
        return {}
    text = open(MD, encoding="utf-8").read()
    sections = re.split(r"^### ", text, flags=re.M)
    stock = {}
    for sec in sections[1:]:
        idx_name = re.sub(r"（.*?）", "", sec.split("\n")[0].strip()).strip()
        if not idx_name:
            continue
        for m in re.finditer(r"\| (\d{6}) \| ([^|]+?) \| ([\d.]+|—) \|", sec):
            code, name, w = m.group(1), m.group(2).strip(), m.group(3)
            s = stock.setdefault(code, {"name": name, "w": {}, "n": 0})
            s["w"][idx_name] = float(w) if w != "—" else None
            s["n"] += 1
    return stock

# ── 2. 行业（东财 push2delay，增量）───────────────────────────────────
def fetch_industry(codes):
    stock = json.load(open(SUMMARY_JSON, encoding="utf-8"))
    todo = [c for c in codes if not stock.get(c, {}).get("industry")]
    print(f"  [行业] 待补 {len(todo)} 只（缓存已有 {len(codes) - len(todo)}）")
    fail = 0
    for i, c in enumerate(todo):
        market = "1" if c.startswith("6") else "0"
        url = ("https://push2delay.eastmoney.com/api/qt/stock/get?fltt=2&invt=2"
               f"&fields=f57,f58,f43,f127&secid={market}.{c}")
        try:
            d = json.loads(em_get(url)).get("data") or {}
            if d.get("f127"):
                stock[c]["industry"] = d["f127"]
                stock[c]["price"] = d.get("f43")
                fail = 0
            else:
                print(f"    [{i}] {c} 空行业，跳过")
        except Exception as e:
            fail += 1
            print(f"    [{i}] {c} 失败({fail}连败): {type(e).__name__} {str(e)[:50]}")
            if fail >= FAIL_LIMIT:
                print("    !!! 连续失败过多，疑似被封，行业补齐中止")
                break
    atomic_dump(SUMMARY_JSON, stock)
    print(f"  [行业] 完成")

# ── 3. 行情（腾讯批量，全量刷新，不封IP）───────────────────────────────
def fetch_quotes(codes):
    def prefix(c):
        if c.startswith("92"):            return "bj" + c   # 北交所必须先于 9x
        if c.startswith(("5", "6", "9")): return "sh" + c
        if c.startswith(("4", "8")):      return "bj" + c
        return "sz" + c
    stock = json.load(open(SUMMARY_JSON, encoding="utf-8"))
    got = 0
    for i in range(0, len(codes), 50):
        batch = codes[i:i + 50]
        url = "https://qt.gtimg.cn/q=" + ",".join(prefix(c) for c in batch)
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            data = urllib.request.urlopen(req, timeout=10).read().decode("gbk", errors="replace")
            for line in data.strip().split(";"):
                if '"' not in line:
                    continue
                key = line.split("=")[0].split("_")[-1]
                v = line.split('"')[1].split("~")
                if len(v) < 53:
                    continue
                code = key[2:]
                if code in stock:
                    stock[code]["t_price"] = float(v[3] or 0)
                    stock[code]["t_pe"] = float(v[39] or 0)
                    stock[code]["t_pb"] = float(v[46] or 0)
                    stock[code]["t_mcap"] = float(v[45] or 0)
                    stock[code]["change_pct"] = float(v[32] or 0)
                    got += 1
        except Exception as e:
            print(f"  [行情] 批次失败: {e}")
        time.sleep(0.3)
    atomic_dump(SUMMARY_JSON, stock)
    print(f"  [行情] 腾讯批量刷新 {got}/{len(codes)}")

# ── 4. 股息率/分红（datacenter，增量；force 全量）─────────────────────
def fetch_dividends(codes, force=False):
    stock = json.load(open(SUMMARY_JSON, encoding="utf-8"))
    todo = [c for c in codes if force or "div_yield_calc" not in stock.get(c, {})]
    if not todo:
        print("  [分红] 缓存已齐全，跳过（--force 可全量重算）")
        return
    print(f"  [分红] 待算 {len(todo)} 只（每只约1.2s，预计 {len(todo) * 1.2 / 60:.1f} 分钟）")
    today = datetime.date.today()
    fail = 0
    done = 0
    for i, c in enumerate(todo):
        url = ("https://datacenter-web.eastmoney.com/api/data/v1/get?"
               "reportName=RPT_SHAREBONUS_DET&columns=ALL"
               f"&filter=(SECURITY_CODE%3D%22{c}%22)"
               "&pageNumber=1&pageSize=8&sortColumns=EX_DIVIDEND_DATE&sortTypes=-1&source=WEB&client=WEB")
        rec = []
        try:
            d = json.loads(em_get(url))
            rows = (d.get("result") or {}).get("data") or []
            for r in rows:
                exdate = str(r.get("EX_DIVIDEND_DATE") or "")[:10]
                bonus = r.get("PRETAX_BONUS_RMB") or 0
                if exdate and bonus:
                    rec.append((exdate, round(bonus, 3)))
            total12 = 0.0
            for exdate, bonus10 in rec:
                try:
                    y, m, _ = map(int, exdate.split("-"))
                    if (today.year - y) * 12 + (today.month - m) <= 12:
                        total12 += bonus10 / 10.0
                except Exception:
                    pass
            price = stock[c].get("t_price") or stock[c].get("price") or 0
            stock[c]["div_yield_calc"] = round(total12 / price * 100, 2) if price and total12 else 0.0
            stock[c]["div_rec"] = rec[:5]
            fail = 0
            done += 1
        except Exception as e:
            fail += 1
            print(f"    [{i}] {c} 失败({fail}连败): {type(e).__name__} {str(e)[:50]}")
            stock[c]["div_yield_calc"] = None
            stock[c]["div_rec"] = []
            if fail >= FAIL_LIMIT:
                print("    !!! 连续失败过多，疑似被封，分红补齐中止")
                break
        if (i + 1) % 50 == 0:
            atomic_dump(SUMMARY_JSON, stock)
            print(f"    进度 {i + 1}/{len(todo)}，成功 {done}，已存盘")
    atomic_dump(SUMMARY_JSON, stock)
    print(f"  [分红] 完成，成功 {done}/{len(todo)}")

# ── 5. 行业归并 + 汇总表缓存 ──────────────────────────────────────────
def build_table():
    sys.path.insert(0, os.path.join(BASE, "scripts"))
    from _classify import map_ind
    stock = json.load(open(SUMMARY_JSON, encoding="utf-8"))
    rows = []
    for c, s in stock.items():
        maxw = max([w for w in s["w"].values() if w], default=0)
        idx = sorted([(k, v) for k, v in s["w"].items() if v], key=lambda x: -x[1])[:8]
        if not idx:   # 权重全部未公开的（如932368/000151成分）
            idx = [[k, None] for k in s["w"].keys()]
        rows.append({
            "code": c, "name": s["name"], "ind3": s.get("industry", ""),
            "ind": map_ind(s.get("industry", "")), "n": s["n"], "maxw": round(maxw, 2),
            "pe": s.get("t_pe"), "pb": s.get("t_pb"),
            "mcap": round(s.get("t_mcap") or 0, 0),
            "price": s.get("t_price") or s.get("price"),
            "div_yield": s.get("div_yield_calc"), "div_rec": s.get("div_rec", []),
            "idx": idx,
        })
    rows.sort(key=lambda r: (r["ind"], -r["n"], -r["maxw"]))
    atomic_dump(TABLE_JSON, rows)
    print(f"  [归并] 汇总表缓存 {len(rows)} 只（{len(set(r['ind'] for r in rows))} 个一级行业）")

# ── 6. 生成 excel/红利成分股汇总.xlsx ─────────────────────────────────
def gen_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    table = json.load(open(TABLE_JSON, encoding="utf-8"))
    stock = json.load(open(SUMMARY_JSON, encoding="utf-8"))
    chg = {c: s.get("change_pct") for c, s in stock.items()}
    wb = Workbook()
    ws = wb.active
    ws.title = "成分股汇总"
    headers = ["序号", "证券代码", "证券名称", "一级行业", "细分行业", "入选指数/ETF数", "最大权重(%)",
               "最新价(元)", "当日涨跌幅(%)", "PE(TTM)", "PB", "总市值(亿元)", "近12个月股息率(%)",
               "近5次分红记录", "主要入选指数及权重(%)", "是否入选20只推荐"]
    head_fill = PatternFill("solid", fgColor="C00000")
    head_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    rec_fill = PatternFill("solid", fgColor="FFF2CC")
    alt_fill = PatternFill("solid", fgColor="F7F7F7")
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.fill, cell.font, cell.alignment, cell.border = head_fill, head_font, center, border
    for i, r in enumerate(table, 1):
        is_rec = r["code"] in FINAL20
        rec_str = "；".join(f"{ex}派{b}元/10股" for ex, b in r["div_rec"][:5]) if r["div_rec"] else "近12个月无派息记录"
        idx_str = "；".join(f"{k} {v}%" if v is not None else f"{k}(权重未公开)" for k, v in r["idx"][:8])
        vals = [i, r["code"], r["name"], r["ind"], r["ind3"], r["n"], r["maxw"],
                r["price"], chg.get(r["code"]), r["pe"], r["pb"], r["mcap"], r["div_yield"],
                rec_str, idx_str, "是" if is_rec else "否"]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=i + 1, column=j, value=v)
            cell.border = border
            cell.alignment = left if j in (2, 3, 14, 15) else center
            if is_rec:
                cell.fill = rec_fill
            elif i % 2 == 0:
                cell.fill = alt_fill
        for j in (8, 9, 10, 11, 13):
            cell = ws.cell(row=i + 1, column=j)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"
        cell = ws.cell(row=i + 1, column=12)
        if isinstance(cell.value, (int, float)):
            cell.number_format = "#,##0"
    for j, w in enumerate([6, 10, 12, 10, 12, 12, 10, 10, 12, 9, 8, 12, 13, 46, 60, 12], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:P{len(table) + 1}"
    # 说明 sheet
    ws2 = wb.create_sheet("说明")
    notes = [
        "红利成分股汇总表（数据日期：最新更新日）",
        "", "数据来源：",
        "1. 成分与权重：《红利指数与ETF成分股.md》解析（精选指数10只+ETF 11只，权重为官方静态快照）",
        "2. 行业：东财个股接口（申万行业分类），一级行业为归并后的申万大类",
        "3. 行情/估值：腾讯财经公开接口",
        "4. 股息率：东财分红历史接口，近12个月每股派息合计÷最新价（近12个月无派息者记 0.00%）",
        "5. 入选20只推荐：指《红利股票推荐20只.md》最终组合（每行业≤2只）",
        "", "列说明：",
        "· 入选指数/ETF数：该股票出现在精选池（10只指数+11只ETF，ETF与跟踪指数成分一致）中的个数，最大21",
        "· 最大权重(%)：该股票在所有入选指数/ETF中的最高权重",
        "· 近5次分红记录：格式为【除权日+每10股税前派息(元)】",
        "· 主要入选指数及权重：按权重降序的前8条，格式【指数名 权重%】；权重未公开（如中证800自由现金流/上证国有企业红利）标注【指数名(权重未公开)】",
        "· PE(TTM) 为负：该股近12个月处于亏损状态，为真实数据",
        "", "风险提示：以上内容仅为基于公开数据的客观信息梳理，不构成投资建议。",
    ]
    for i, line in enumerate(notes, 1):
        ws2.cell(row=i, column=1, value=line)
    ws2.column_dimensions["A"].width = 110
    out = os.path.join(EXCEL, "红利成分股汇总.xlsx")
    wb.save(out)
    print(f"  [Excel] 已生成 {out}（{len(table)} 只 × {len(headers)} 列）")

def stock_change(code):
    stock = json.load(open(SUMMARY_JSON, encoding="utf-8"))
    return stock.get(code, {}).get("change_pct")
# ── 7. 核对推荐20只 ───────────────────────────────────────────────────
def check_final20():
    stock = json.load(open(SUMMARY_JSON, encoding="utf-8"))
    missing = [c for c in FINAL20 if c not in stock]
    print(f"  [核对] 推荐20只：{len(FINAL20) - len(missing)} 只在当前成分中，{len(missing)} 只不在")
    if missing:
        print("  ⚠️ 以下股票已不在最新成分股中，请人工核对《红利股票推荐20只.md》：", missing)
    else:
        print("  ✅ 20只全部仍在成分股中")

# ── 主流程 ────────────────────────────────────────────────────────────
def run(force=False):
    t0 = time.time()
    print("═══ 成分股汇总数据更新 ═══")
    parsed = parse_components()
    if not parsed:
        print("  ❌ 无成分股数据可处理，终止本次更新（缓存未做任何改动）")
        return
    print(f"[1/7] 解析成分股md：{len(parsed)} 只")
    old = {}
    if os.path.exists(SUMMARY_JSON):
        try:
            old = json.load(open(SUMMARY_JSON, encoding="utf-8"))
        except Exception as e:
            print(f"  ⚠️ 缓存 {SUMMARY_JSON} 损坏（{type(e).__name__}），将基于 md 重建")
    if old:
        new_codes = [c for c in parsed if c not in old]
        if new_codes:
            print(f"  md 新增 {len(new_codes)} 只：{new_codes}")
        for c, s in parsed.items():
            if c not in old:
                old[c] = s
        # 同步：md 中已移除的股票从缓存剔除（带保护：md 解析异常/骤降时禁止清空）
        removed = [c for c in old if c not in parsed]
        if removed:
            if len(removed) > len(old) // 2:
                print(f"  ⚠️ 预警：md 解析出 {len(parsed)} 只，但旧缓存有 {len(old)} 只，其中 {len(removed)} 只不在新解析结果中")
                print(f"     （疑似 md 损坏或解析异常），跳过删除同步，仅保留新增；请人工核对 md 后再更新")
            else:
                print(f"  md 已移除 {len(removed)} 只：{removed}，从缓存删除")
                for c in removed:
                    del old[c]
        atomic_dump(SUMMARY_JSON, old)
        print("[2/7] 缓存基础信息已同步")
    else:
        for c, s in parsed.items():
            old[c] = s
        atomic_dump(SUMMARY_JSON, old)
        print("[2/7] 首次生成基础缓存（或缓存损坏后基于 md 重建）")
    codes = list(parsed.keys())
    print("[3/7] 行业补齐（增量）")
    fetch_industry(codes)
    print("[4/7] 行情刷新（腾讯批量）")
    fetch_quotes(codes)
    print("[5/7] 股息率/分红")
    fetch_dividends(codes, force=force)
    print("[6/7] 行业归并 + 汇总表缓存")
    build_table()
    print("[7/7] 生成 Excel")
    gen_excel()
    check_final20()
    print(f"✅ 完成，耗时 {(time.time() - t0) / 60:.1f} 分钟")
    print("⚠️ 请人工核对《红利股票推荐20只.md》中相关数字/描述是否需要同步（脚本不自动改该文件）。")

if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    run(force="--force" in sys.argv)
