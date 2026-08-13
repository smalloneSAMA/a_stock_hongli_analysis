# -*- coding: utf-8 -*-
"""自选股（excel/自选股清单.xlsx）历史数据拉取 / 增量更新

清单：excel/自选股清单.xlsx 为唯一事实来源（表头：序号/股票名/股票代码），每次运行现读，
      代码列可能是数字存储（如 601899）→ zfill(6) 防丢前导零。
数据：与 _fetch_pool_data 完全同构 —— K线（腾讯，增量）+ 分红/财报/股本（东财，缺失补齐）。
      清单内股票若已在精选池（已有缓存），自动复用零成本。
行业：东财 push2delay 增量补（仅新股票），写 cache/_自选股清单.json
      （供 manifest 展示；该文件仅存增强信息，股票列表始终以 xlsx 为准）。
防封：节流 0.8s/只、失败清单 cache/_watchlist_failed.json（--retry-failed 只补失败）。

用法:
  python scripts/_fetch_watchlist.py               # 增量：K线增量 + 缺失补齐（分钟级）
  python scripts/_fetch_watchlist.py --indicators  # 全量刷新清单全部股票的 总市值/PE-TTM/PB（腾讯批量）→ 写回 xlsx E/F/G 列
  python scripts/_fetch_watchlist.py --retry-failed # 只重试失败清单
  python scripts/_fetch_watchlist.py --check-fin   # 季度分红/财报/股本检测（约 2s/只）
"""
import sys, os, json, time, argparse, urllib.request

sys.stdout.reconfigure(encoding="utf-8")
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(BASE, "scripts"))

import _fetch_history as fh
import _fetch_stock_data as fsd
from _common import market_prefix, tencent_quotes
from _classify import map_ind

XLSX = os.path.join(BASE, "excel", "自选股清单.xlsx")   # 自选股清单（唯一事实来源）
META_PATH = os.path.join(BASE, "cache", "_自选股清单.json")   # 仅行业等增强信息
METRICS_PATH = os.path.join(BASE, "cache", "_自选股指标.json")  # 指标缓存（总市值/PE-TTM/PB + 数据日期）
FAIL_PATH = os.path.join(BASE, "cache", "_watchlist_failed.json")
SLEEP_KLINE = 0.8      # 腾讯 K 线节流（秒）
MELT_LIMIT = 5         # 连续失败熔断阈值
MELT_PAUSE = 60        # 熔断暂停（秒）


def read_watchlist_xlsx():
    """现读 xlsx → [{code, name, tcode, show, seq}]
    表头：序号/股票名/股票代码/展示（第4列，==1 才展示，其余不展示）；跳过空行/非法行。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  ❌ 缺少 openpyxl（pip install openpyxl）")
        return []
    if not os.path.exists(XLSX):
        print(f"  ⚠️ 找不到自选股清单（{XLSX}）")
        return []
    wb = load_workbook(XLSX, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    out = []
    seen = set()
    for row in ws.iter_rows(values_only=True):
        if row is None or len(row) < 3:
            continue
        name, raw = row[1], row[2]
        if name is None or raw is None:
            continue
        name = str(name).strip()
        code = str(raw).strip()
        if code.endswith(".0"):        # openpyxl 可能读出 float
            code = code[:-2]
        if not code.isdigit() or len(code) > 6:
            continue                   # 表头行/脏数据
        code = code.zfill(6)
        if code in seen:
            continue
        seen.add(code)
        show_v = row[3] if len(row) > 3 else None
        show = show_v == 1 or (isinstance(show_v, str) and show_v.strip() == "1")
        seq_v = row[0]
        seq = int(seq_v) if isinstance(seq_v, (int, float)) and seq_v > 0 else 0
        out.append({"code": code, "name": name, "tcode": market_prefix(code),
                    "show": show, "seq": seq})
    wb.close()
    if not out:
        print("  ⚠️ 自选股清单为空或解析失败（表头应为：序号/股票名/股票代码/展示）")
    return out


def load_watch_meta():
    """读行业增强缓存（供 _gen_web_data 展示用；无则返回 {}，不影响主流程）"""
    try:
        return json.load(open(META_PATH, encoding="utf-8")).get("rows", {})
    except OSError:
        return {}


def refresh_watch_meta(rows):
    """增量补行业（东财 push2delay f127，仅新股票；复用 fsd.em_get 限流）→ 写缓存"""
    meta = load_watch_meta()
    todo = [r for r in rows if r["show"] and not meta.get(r["code"], {}).get("ind3")]
    if not todo:
        return
    print(f"── 补行业（东财，{len(todo)} 只）──")
    fail = 0
    for i, r in enumerate(todo, 1):
        market = "1" if r["code"].startswith("6") else "0"
        url = ("https://push2delay.eastmoney.com/api/qt/stock/get?fltt=2&invt=2"
               f"&fields=f57,f58,f43,f127&secid={market}.{r['code']}")
        try:
            d = json.loads(fsd.em_get(url)).get("data") or {}
            ind3 = d.get("f127") or ""
            if ind3:
                meta[r["code"]] = {"ind3": ind3, "ind": map_ind(ind3)}
                fail = 0
                print(f"  [{i}/{len(todo)}] {r['code']} {r['name']} → {map_ind(ind3)} / {ind3}")
            else:
                print(f"  [{i}/{len(todo)}] {r['code']} {r['name']} 空行业，跳过")
        except Exception as e:
            fail += 1
            print(f"  [{i}/{len(todo)}] {r['code']} 行业失败: {repr(e)[:60]}")
            if fail >= 8:
                print("  !!! 连续失败过多，疑似被封，行业补齐中止")
                break
    tmp = META_PATH + ".tmp"
    with open(tmp, "w", encoding="utf-8") as fp:
        json.dump({"updated_at": time.strftime("%Y-%m-%d"), "rows": meta}, fp, ensure_ascii=False, indent=1)
    os.replace(tmp, META_PATH)
    print(f"  ✅ 行业缓存已更新 → {META_PATH}")


def refresh_indicators_main():
    """全量刷新清单**全部**股票的 总市值(亿)/PE(TTM)/PB（腾讯批量 50只/批，不封IP）
    → 缓存 cache/_自选股指标.json → 写回 xlsx 第 5/6/7 列（表头第 1 行，数据按 C 列代码逐行匹配，其余列不碰）。
    口径与成分股汇总一致：腾讯 v[45]=总市值(元)、v[39]=PE(TTM)、v[46]=PB。"""
    rows = read_watchlist_xlsx()
    if not rows:
        return
    print(f"═══ 自选股清单指标刷新（{len(rows)} 只 · 腾讯批量 · 全量）═══")

    metrics, got, t0 = {}, 0, time.time()
    for code, v in tencent_quotes([r["code"] for r in rows]).items():
        price, pe, pb = float(v[3] or 0), float(v[39] or 0), float(v[46] or 0)
        mcap = float(v[45] or 0)   # 腾讯总市值，单位已是亿元（招行 9785.30 = 9785 亿）
        if not (mcap or pe or pb):
            continue   # 无行情（停牌/退市/代码错误）→ 不写缓存
        metrics[code] = {"t_price": round(price, 2),
                         "t_pe": round(pe, 2) if pe else None,
                         "t_pb": round(pb, 2) if pb else None,
                         "t_mcap": round(mcap, 2) if mcap else None}
        got += 1
    print(f"  腾讯批量行情完成，命中 {got}/{len(rows)}（{time.time() - t0:.0f}s）")
    tmp = METRICS_PATH + ".tmp"
    with open(tmp, "w", encoding="utf-8") as fp:
        json.dump({"date": time.strftime("%Y-%m-%d"), "rows": metrics}, fp, ensure_ascii=False, indent=1)
    os.replace(tmp, METRICS_PATH)
    print(f"  ✅ 指标缓存已更新（{got} 只）→ {METRICS_PATH}")
    # 写回 xlsx（先落盘缓存，写 xlsx 失败可重跑不重拉）
    try:
        write_indicators_xlsx(metrics)
        print("  ✅ 已写回 excel/自选股清单.xlsx（追加列：总市值(亿)/PE(TTM)/PB，已有列未动）")
    except PermissionError:
        print("  ⚠️ xlsx 被占用（Excel 中打开？）请关闭后重跑 --indicators（缓存已存，重跑约30秒不重拉）")
    except Exception as e:
        print(f"  ⚠️ 写回 xlsx 失败: {repr(e)[:80]}（缓存已存，可重跑）")


def write_indicators_xlsx(metrics):
    """写回 xlsx：新三列（总市值(亿)/PE(TTM)/PB）**追加到现有表头之后**（H/I/J…），已有列一律不碰；
    若表头已存在同名列则原位覆盖（幂等，重复运行不产生新列）。
    表头在第 1 行，数据行按 C 列代码逐行匹配。"""
    from openpyxl import load_workbook
    wb = load_workbook(XLSX)
    ws = wb.worksheets[0]
    # 探测表头：第 1 行全部列名 → 定位三列（已有则原位，否则追加）
    headers, last_col = {}, 0
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is not None:
            headers[str(v).strip()] = c
            last_col = c
    order = ["总市值(亿)", "PE(TTM)", "PB"]
    pos = {}
    for k in order:
        if k in headers:
            pos[k] = headers[k]
        else:
            last_col += 1
            ws.cell(1, last_col, k)
            pos[k] = last_col
    n = 0
    for r in range(2, ws.max_row + 1):
        raw = ws.cell(r, 3).value
        if raw is None:
            continue
        code = str(raw).strip()
        if code.endswith(".0"):
            code = code[:-2]
        m = metrics.get(code.zfill(6))
        if not m:
            continue
        ws.cell(r, pos["总市值(亿)"], m["t_mcap"]); ws.cell(r, pos["PE(TTM)"], m["t_pe"]); ws.cell(r, pos["PB"], m["t_pb"])
        n += 1
    wb.save(XLSX)
    wb.close()
    print(f"  写回 {n} 行（命中率 {n / max(ws.max_row - 1, 1) * 100:.1f}%）")


def load_failed():
    try:
        return json.load(open(FAIL_PATH, encoding="utf-8"))
    except OSError:
        return {}


def save_failed(f):
    tmp = FAIL_PATH + ".tmp"
    with open(tmp, "w", encoding="utf-8") as fp:
        json.dump(f, fp, ensure_ascii=False, indent=1)
    os.replace(tmp, FAIL_PATH)


def fetch_one(code, name, tcode):
    """单只：K线增量 + 分红/财报/股本缺失补齐。抛异常表示 K 线失败（入失败清单）"""
    fh.update_incremental("股票", code, name, fsd.make_fetcher(tcode, code))
    fsd.update_dividends(codes=[code])
    fsd.update_financials(codes=[code])
    fsd.update_share_hist(codes=[code])


def main(retry_failed=False, check_fin=False):
    rows = read_watchlist_xlsx()
    if not rows:
        return
    rows = [r for r in rows if r["show"]]   # 仅展示字段==1 的股票参与拉取/展示
    if not rows:
        print("⚠️ 自选股清单无展示股票（展示字段需为 1）")
        return
    if check_fin:
        print(f"═══ 自选股分红/财报/股本季度检测（{len(rows)} 只）═══")
        fsd.check_financials(codes=[r["code"] for r in rows])
        fsd.update_share_hist(codes=[r["code"] for r in rows])
        print("✅ 季度检测完成（随后运行 python scripts/_gen_web_data.py 重算指标）")
        return

    failed = load_failed()
    todo = rows
    if retry_failed:
        fset = set(failed)
        todo = [r for r in rows if r["code"] in fset]
        if not todo:
            print("✅ 失败清单无待重试项（cache/_watchlist_failed.json 已清空或全成功）")
            save_failed({})
            return
        print(f"═══ 重试失败清单（{len(todo)} 只）═══")
    else:
        print(f"═══ 自选股增量更新（{len(rows)} 只：{', '.join(r['name'] for r in rows)}）═══")

    # 行业增强信息（仅新股票，失败不影响行情主流程）
    refresh_watch_meta(rows)

    new_failed = {k: v for k, v in failed.items()}
    consec = 0
    ok = 0
    t0 = time.time()
    for i, r in enumerate(todo, 1):
        try:
            fetch_one(r["code"], r["name"], r["tcode"])
            new_failed.pop(r["code"], None)
            consec = 0
            ok += 1
            if i % 10 == 0 or i == len(todo):
                print(f"  进度 {i}/{len(todo)}（成功 {ok}） 已用 {time.time() - t0:.0f}s")
        except Exception as e:
            consec += 1
            new_failed[r["code"]] = {"name": r["name"], "err": f"{type(e).__name__}: {str(e)[:70]}"}
            print(f"  ❌ [{i}/{len(todo)}] {r['code']} {r['name']} 拉取失败: {repr(e)[:80]}")
            if consec >= MELT_LIMIT:
                print(f"  ⏸ 连续 {MELT_LIMIT} 次失败，熔断暂停 {MELT_PAUSE}s 防封…")
                time.sleep(MELT_PAUSE)
                consec = 0
        time.sleep(SLEEP_KLINE)
    save_failed(new_failed)
    fail_n = len(todo) - ok
    print(f"✅ 完成：成功 {ok} / 失败 {fail_n}（失败明细 → cache/_watchlist_failed.json）")
    if fail_n:
        print("   下次运行 --retry-failed 只补失败项")
    print("⚠️  随后请运行 update.py 选项7（前端数据包 + 区间分析）")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--indicators", action="store_true", help="全量刷新清单指标（总市值/PE-TTM/PB，腾讯批量）并写回 xlsx")
    ap.add_argument("--retry-failed", action="store_true", help="只重试失败清单")
    ap.add_argument("--check-fin", action="store_true", help="季度分红/财报/股本检测")
    args = ap.parse_args()
    if args.indicators:
        refresh_indicators_main()
    else:
        main(retry_failed=args.retry_failed, check_fin=args.check_fin)
