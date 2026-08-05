# -*- coding: utf-8 -*-
"""
拉取国证指数完整成分（国证官网 sample-detail/detail 接口，纯 HTTP，无需浏览器）

- 目前用于 980092 国证自由现金流（100 只）；接口通用，indexcode 参数化
- 字段：seccode/secname/trade(国证行业)/weight(权重)；官网仅披露前十大权重，其余为 "--"
- 输出：cache/成分_{code}.json（原子写）：{code, name, sample_date, total, top10, industry_dist, stocks}
- 频率：季度调样后重跑一次即可（低频免费接口）

用法: python scripts/_fetch_cnindex_components.py [indexcode ...]
"""
import io, json, os, sys, time, urllib.request

if __name__ == "__main__":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
URL = "https://www.cnindex.com.cn/sample-detail/detail"
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"

# 默认清单：980092 国证自由现金流；可用命令行参数追加/覆盖
INDICES = [("980092", "国证自由现金流")]


def atomic_dump(path, obj):
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False)
    os.replace(tmp, path)


def fetch(code, name, date_str=None):
    date_str = date_str or time.strftime("%Y-%m")
    url = f"{URL}?indexcode={code}&dateStr={date_str}&pageNum=1&rows=100&isFirstCall=1"
    req = urllib.request.Request(url, headers={
        "User-Agent": UA,
        "Referer": f"https://www.cnindex.com.cn/module/index-detail.html?act_menu=1&indexCode={code}"})
    d = json.loads(urllib.request.urlopen(req, timeout=15).read().decode("utf-8"))
    rows = (d.get("data") or {}).get("rows") or []
    if not rows:
        raise ValueError(f"接口返回空（{url}）")
    stocks = []
    for r in rows:
        w = r.get("weight")
        stocks.append({"code": str(r.get("seccode") or "").zfill(6),
                       "name": r.get("secname") or "",
                       "ind": r.get("trade") or "",
                       "weight": None if w in (None, "--", "") else float(w)})
    industry_dist = {}
    for s in stocks:
        industry_dist[s["ind"]] = industry_dist.get(s["ind"], 0) + 1
    top10 = [s for s in stocks if s["weight"] is not None][:10]
    return {"code": code, "name": name,
            "sample_date": (rows[0].get("dateStr") or "")[:10],
            "total": d.get("total") or len(rows),
            "stocks": stocks, "top10": top10,
            "industry_dist": industry_dist}


def export_excel(objs):
    """生成 excel/国证指数成分.xlsx：每只国证指数一个 sheet（代码/名称/国证行业/权重），附说明页"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    wb.remove(wb.active)
    head_fill = PatternFill("solid", fgColor="C00000")
    head_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    for obj in objs:
        ws = wb.create_sheet(obj["code"])
        headers = ["序号", "证券代码", "证券名称", "国证行业", "权重(%)"]
        for j, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=j, value=h)
            cell.fill, cell.font, cell.alignment, cell.border = head_fill, head_font, center, border
        for i, s in enumerate(obj["stocks"], 1):
            w = "—" if s["weight"] is None else round(s["weight"], 2)
            vals = [i, s["code"], s["name"], s["ind"], w]
            for j, v in enumerate(vals, 1):
                cell = ws.cell(row=i + 1, column=j, value=v)
                cell.border, cell.alignment = border, center
        for j, wd in enumerate([6, 10, 12, 12, 10], 1):
            ws.column_dimensions[get_column_letter(j)].width = wd
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:E{len(obj['stocks']) + 1}"
    ws2 = wb.create_sheet("说明")
    notes = ["国证指数成分股（数据源：国证指数官网样本详情接口 sample-detail/detail）", "", "口径：",
             "· 官网仅披露前十大权重，其余成分权重记 —（未公开）",
             "· 行业为国证行业分类（与申万分类口径不同）",
             "· 样本随季度调样更新，重新运行 scripts/_fetch_cnindex_components.py 刷新", "",
             "风险提示：以上内容仅为基于公开数据的客观信息梳理，不构成投资建议。"]
    for i, line in enumerate(notes, 1):
        ws2.cell(row=i, column=1, value=line)
    ws2.column_dimensions["A"].width = 90
    out = os.path.join(BASE, "excel", "国证指数成分.xlsx")
    wb.save(out)
    print(f"✅ excel/国证指数成分.xlsx 已生成（{len(objs)} 个指数 sheet）")


def main():
    codes = sys.argv[1:] or [c for c, _ in INDICES]
    objs = []
    for code in codes:
        name = dict(INDICES).get(code, code)
        ok = False
        for attempt in range(3):
            try:
                obj = fetch(code, name)
                p = os.path.join(BASE, "cache", f"成分_{code}.json")
                atomic_dump(p, obj)
                w10 = sum((s["weight"] or 0) for s in obj["top10"])
                print(f"✅ {code} {name}: {obj['total']} 只，样本日期 {obj['sample_date']}，"
                      f"前十大权重 {len(obj['top10'])} 只合计 {w10:.2f}%，行业 {len(obj['industry_dist'])} 个")
                print(f"   → {p}")
                objs.append(obj)
                ok = True
                break
            except Exception as e:
                print(f"  ⚠️ [{code}] 第{attempt + 1}次失败: {type(e).__name__} {str(e)[:60]}")
                time.sleep(2 + attempt * 2)
        if not ok:
            print(f"  ❌ [{code}] 拉取失败")
    if objs:
        export_excel(objs)


if __name__ == "__main__":
    main()
