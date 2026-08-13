# -*- coding: utf-8 -*-
"""
推荐20只股票历史日线拉取工具（不复权，2004-01-01 起）

- 标的：《红利股票推荐20只.md》全部20只
- 数据源：腾讯日K（不复权，真实历史价格；前复权早期价格会因分红变负，故不用）
- 全量：翻页拉取，过滤 2004-01-01 之前数据（上市晚于该日的自然从上市日起）
- 增量：腾讯 start 参数实测无效（返回最近800条），增量=拉最近800条+字段级合并覆盖，
        等价于刷新最近约3.2年，更早数据保留
- 分红：东财 RPT_SHAREBONUS_DET 全量分红历史 → cache/分红_{code}.json（原子写，缓存存在跳过）
- 股息率：Excel 每行 = 除权日在(当日-365天,当日]内每股派息 ÷ 当日收盘 ×100（精确自然日窗口，同东财口径）
- 财报：东财 RPT_F10_FINANCE_MAINFINADATA → cache/财报_{code}.json（累计归母净利 PARENTNETPROFIT + EPS/BPS/ROE/ROA）
- 股本：东财 RPT_F10_EH_EQUITY 股本变动历史 → cache/股本_{code}.json（按变动生效日阶梯取当日总股本）
- PE：总市值(收盘×当日总股本) ÷ 归母净利（TTM/年化），口径同东财行情页；不用"价÷EPS"避免加权股本≠期末股本偏差
- 缓存：cache/股票_{code}.json，原子写（_fetch_history.save_cache），无缓存全量自愈
- 成交额：腾讯日K无成交额字段，按 成交量(手)×100×(高+低+收)/3 估算（同ETF口径）
- Excel：excel/股票历史.xlsx（每只一个sheet，含股息率(%)列）
用法: python scripts/_fetch_stock_data.py [--refresh]
"""
import sys, io, os, json, time, argparse, urllib.request, random
import pandas as pd

if __name__ == "__main__":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # 项目根
sys.path.insert(0, os.path.join(BASE, "scripts"))
import _fetch_history as fh
from _common import em_get   # 东财限流请求（1s/请求防封，全局限流）

# ── 推荐股清单（动态：读 cache/_推荐20.json 评分产物；缺失时回退硬编码清单）──
# 推荐清单由 scripts/_recommend_stocks.py 生成（量化评分）；此清单同时驱动 manifest rec 标记、
# 回测“推荐20”分组、推荐股 Excel 导出与 K 线增量拉取——单一来源，全链路自动跟随
_FALLBACK_STOCKS = [
    ("600036", "招商银行", "sh600036"),
    ("601838", "成都银行", "sh601838"),
    ("601088", "中国神华", "sh601088"),
    ("601225", "陕西煤业", "sh601225"),
    ("600938", "中国海油", "sh600938"),
    ("601857", "中国石油", "sh601857"),
    ("600350", "山东高速", "sh600350"),
    ("601006", "大秦铁路", "sh601006"),
    ("600900", "长江电力", "sh600900"),
    ("600795", "国电电力", "sh600795"),
    ("000858", "五粮液", "sz000858"),
    ("000895", "双汇发展", "sz000895"),
    ("000651", "格力电器", "sz000651"),
    ("000333", "美的集团", "sz000333"),
    ("000423", "东阿阿胶", "sz000423"),
    ("600566", "济川药业", "sh600566"),
    ("600019", "宝钢股份", "sh600019"),
    ("601668", "中国建筑", "sh601668"),
    ("600582", "天地科技", "sh600582"),
    ("600757", "长江传媒", "sh600757"),
]


def _rec_stocks():
    """读评分产物 → [(code, name, tcode)]；产物缺失/损坏回退硬编码清单"""
    p = os.path.join(BASE, "cache", "_推荐20.json")
    try:
        obj = json.load(open(p, encoding="utf-8"))
        rows = obj.get("list") or []
        if rows:
            out = []
            for r in rows:
                code = r.get("code")
                if not code:
                    continue
                tcode = ("sh" if code.startswith(("6", "9")) else "sz") + code
                out.append((code, r.get("name", code), tcode))
            if out:
                return out
    except Exception:
        pass
    return list(_FALLBACK_STOCKS)


STOCKS = _rec_stocks()


def refresh_stocks():
    """评分产物更新后刷新进程内 STOCKS（模块 import 时求值一次，产物驱动需手动重读）"""
    global STOCKS
    STOCKS = _rec_stocks()


START = "2004-01-01"  # 最早时间上限
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/126.0 Safari/537.36"

# 东财限流请求统一走 _common.em_get（1s/请求防封，全局限流）

# ── 全量分红历史（东财 RPT_SHAREBONUS_DET，每10股税前派息）──────────
def fetch_dividend(code):
    """返回 [{ex_date, bonus10}]，按除权日降序（接口默认）；"""
    url = ("https://datacenter-web.eastmoney.com/api/data/v1/get?"
           "reportName=RPT_SHAREBONUS_DET&columns=ALL"
           f"&filter=(SECURITY_CODE%3D%22{code}%22)"
           "&pageNumber=1&pageSize=50&sortColumns=EX_DIVIDEND_DATE&sortTypes=-1&source=WEB&client=WEB")
    d = json.loads(em_get(url))
    rows = (d.get("result") or {}).get("data") or []
    out = []
    for r in rows:
        exdate = str(r.get("EX_DIVIDEND_DATE") or "")[:10]
        bonus = r.get("PRETAX_BONUS_RMB") or 0
        if exdate and bonus:
            out.append({"ex_date": exdate, "bonus10": round(bonus, 3)})
    return out

def update_dividends(refresh=False, codes=None):
    """全量分红 → cache/分红_{code}.json（缓存已存在则跳过，删除自愈重拉；refresh=True 强制重拉）
    codes=None 默认推荐20只；传 code 列表可拉其他成份股（_fetch_pool_data）"""
    lst = codes if codes is not None else [c for c, _, _ in STOCKS]
    names = {c: n for c, n, _ in STOCKS}
    print(f"── 全量分红历史（东财，1s/只，{len(lst)} 只）──")
    for code in lst:
        name = names.get(code, code)
        if not refresh and os.path.exists(fh.cache_path("分红", code)):
            print(f"  [{code} {name}] 分红缓存已存在，跳过")
            continue
        try:
            rows = fetch_dividend(code)
            obj = {"code": code, "name": name,
                   "fetched_at": time.strftime("%Y-%m-%d"), "rows": rows}
            fh.save_cache("分红", code, obj)   # 原子写
            print(f"  [{code} {name}] 分红 {len(rows)} 条 -> cache/分红_{code}.json")
        except Exception as e:
            print(f"  ❌ [{code} {name}] 分红拉取失败: {repr(e)[:80]}")
        time.sleep(0.3)

# ── 财报（东财 RPT_F10_FINANCE_MAINFINADATA，季度累计EPS/每股净资产/ROE/ROA）──
def fetch_financials(code):
    """返回 [{report_date, notice_date, eps, np, share, bps, roe, roa}]，按报告期降序；
    EPSJB=基本每股收益(累计)，PARENTNETPROFIT=归母净利润(元,累计)，
    TOTAL_SHARE=总股本(股)——注意该字段全历史回填当前值，仅作股本缓存缺失时的回退；
    BPS=每股净资产，ROEJQ=加权净资产收益率(%)，ZZCJLL=总资产净利率(%)（ROA口径）"""
    secucode = code + (".SH" if code.startswith("6") else ".SZ")
    url = ("https://datacenter-web.eastmoney.com/api/data/v1/get?"
           "reportName=RPT_F10_FINANCE_MAINFINADATA&columns=ALL"
           f"&filter=(SECUCODE%3D%22{secucode}%22)&pageNumber=1&pageSize=200"
           "&sortColumns=REPORT_DATE&sortTypes=-1&source=WEB&client=WEB")
    d = json.loads(em_get(url))
    rows = (d.get("result") or {}).get("data") or []
    out = []
    for r in rows:
        eps = r.get("EPSJB")
        bps = r.get("BPS")
        roe = r.get("ROEJQ")
        roa = r.get("ZZCJLL")
        np = r.get("PARENTNETPROFIT")   # 归母净利润（元，累计）
        ts = r.get("TOTAL_SHARE")       # 总股本（股）——全期回填当前值，仅应急回退用
        rd = str(r.get("REPORT_DATE") or "")[:10]
        nd = str(r.get("NOTICE_DATE") or "")[:10]
        if rd and nd and eps:
            out.append({"report_date": rd, "notice_date": nd, "eps": round(float(eps), 4),
                        "np": round(float(np), 2) if np else None,
                        "share": int(ts) if ts else None,
                        "bps": round(float(bps), 4) if bps else None,
                        "roe": round(float(roe), 4) if roe else None,     # 0=接口缺失标记，视为无值
                        "roa": round(float(roa), 4) if roa else None})
    return out

# ── 股本变动历史（东财 RPT_F10_EH_EQUITY，按变动生效日阶梯）─────────
def fetch_share_history(code):
    """返回 [{date, total_share}]（按生效日升序，总股本=股）。
    END_DATE=股本变动生效日，TOTAL_SHARES=当日总股本；接口失败回退 F10 股本结构页 lngbbd
    注意：送转/回购频繁的股票记录可达上千条（如美的每笔回购一条），需翻页拉全"""
    secucode = code + (".SH" if code.startswith("6") else ".SZ")
    out, seen = [], set()
    try:
        for page in range(1, 21):   # 最多 20页×500=1万条，足够覆盖全部股本变动
            url = ("https://datacenter-web.eastmoney.com/api/data/v1/get?"
                   "reportName=RPT_F10_EH_EQUITY&columns=ALL"
                   f"&filter=(SECUCODE%3D%22{secucode}%22)&pageNumber={page}&pageSize=500"
                   "&sortColumns=END_DATE&sortTypes=-1&source=WEB&client=WEB")
            d = json.loads(em_get(url))
            rows = (d.get("result") or {}).get("data") or []
            if not rows:
                break
            for r in rows:
                dt = str(r.get("END_DATE") or "")[:10]
                ts = r.get("TOTAL_SHARES")
                if dt and ts and dt not in seen:
                    seen.add(dt)
                    out.append({"date": dt, "total_share": int(ts)})
            if len(rows) < 500:
                break
            time.sleep(0.3)
        if out:
            return sorted(out, key=lambda x: x["date"])
    except Exception:
        pass
    # 回退：F10 股本结构页（lngbbd 可能只回最近约20条）
    try:
        market = "SH" if code.startswith("6") else "SZ"
        url2 = (f"https://emweb.securities.eastmoney.com/PC_HSF10/CapitalStockStructure/"
                f"PageAjax?code={market}{code}")
        req = urllib.request.Request(url2, headers={
            "User-Agent": UA, "Referer": "https://emweb.securities.eastmoney.com/"})
        d2 = json.loads(urllib.request.urlopen(req, timeout=12).read().decode())
        out = []
        for r in sorted(d2.get("lngbbd") or [], key=lambda x: str(x.get("END_DATE") or "")):
            dt = str(r.get("END_DATE") or "")[:10]
            ts = r.get("TOTAL_SHARES")
            if dt and ts:
                out.append({"date": dt, "total_share": int(ts)})
        return out
    except Exception:
        return []

def update_share_hist(refresh=False, codes=None):
    """股本变动历史 → cache/股本_{code}.json（缓存存在跳过；refresh=True 强制重拉）"""
    lst = codes if codes is not None else [c for c, _, _ in STOCKS]
    names = {c: n for c, n, _ in STOCKS}
    print(f"── 股本变动历史（东财 EH_EQUITY，1s/只，{len(lst)} 只）──")
    for code in lst:
        name = names.get(code, code)
        if not refresh and os.path.exists(fh.cache_path("股本", code)):
            print(f"  [{code} {name}] 股本缓存已存在，跳过")
            continue
        try:
            rows = fetch_share_history(code)
            obj = {"code": code, "name": name,
                   "fetched_at": time.strftime("%Y-%m-%d"), "rows": rows}
            fh.save_cache("股本", code, obj)
            print(f"  [{code} {name}] 股本变动 {len(rows)} 条 -> cache/股本_{code}.json")
        except Exception as e:
            print(f"  ❌ [{code} {name}] 股本拉取失败: {repr(e)[:80]}")
        time.sleep(0.3)

def update_financials(refresh=False, codes=None):
    """财报 → cache/财报_{code}.json（缓存存在且为新版字段则跳过；旧版缺归母净利自动重拉）
    同时确保股本变动历史缓存存在（update_share_hist）；refresh=True 强制全量重拉"""
    lst = codes if codes is not None else [c for c, _, _ in STOCKS]
    names = {c: n for c, n, _ in STOCKS}
    update_share_hist(refresh=refresh, codes=codes)
    print(f"── 季度财报（东财，1s/只，{len(lst)} 只）──")
    for code in lst:
        name = names.get(code, code)
        stale = refresh
        if os.path.exists(fh.cache_path("财报", code)) and not refresh:
            try:
                old = json.load(open(fh.cache_path("财报", code), encoding="utf-8"))
                if not old.get("rows") or "np" not in (old["rows"][0] or {}):
                    stale = True   # 旧版缓存缺归母净利字段，重拉自愈
            except Exception:
                stale = True
            if not stale:
                print(f"  [{code} {name}] 财报缓存已存在，跳过")
                continue
            print(f"  [{code} {name}] 财报缓存为旧版(缺归母净利)，重拉")
        else:
            print(f"  [{code} {name}] 财报缓存缺失，重拉")
        try:
            rows = fetch_financials(code)
            obj = {"code": code, "name": name,
                   "fetched_at": time.strftime("%Y-%m-%d"), "rows": rows}
            fh.save_cache("财报", code, obj)   # 原子写
            print(f"  [{code} {name}] 财报 {len(rows)} 条 -> cache/财报_{code}.json")
        except Exception as e:
            print(f"  ❌ [{code} {name}] 财报拉取失败: {repr(e)[:80]}")
        time.sleep(0.3)

# ── 历史 PE 计算（市值÷归母净利口径，同东财行情页）───────────────────
def calc_ttm_np(rows, fin_rows):
    """每行 TTM 归母净利润：最新累计归母净利 − 去年同期累计 + 去年年报（累计口径折算12个月）"""
    import bisect
    fins = sorted(fin_rows, key=lambda x: (x["notice_date"], x["report_date"]))
    dates = [f["notice_date"] for f in fins]
    by_yq = {}
    for f in fins:  # 公告晚的覆盖早的（同报告期以最新公告为准）
        by_yq[(f["report_date"][:4], f["report_date"][5:10])] = f
    out = []
    for r in rows:
        d = r["date"]
        i = bisect.bisect_right(dates, d) - 1
        if i < 0:
            out.append(None)
            continue
        latest = fins[i]
        ly, lq = latest["report_date"][:4], latest["report_date"][5:10]
        same_ly = by_yq.get((str(int(ly) - 1), lq))
        annual_ly = by_yq.get((str(int(ly) - 1), "12-31"))
        ttm_np = None
        if latest.get("np") is None:
            pass
        elif same_ly and annual_ly and same_ly.get("np") is not None and annual_ly.get("np") is not None:
            ttm_np = latest["np"] - same_ly["np"] + annual_ly["np"]
        elif lq == "12-31":
            ttm_np = latest["np"]   # 最新即年报
        out.append(ttm_np)
    return out

def calc_share(rows, share_rows, fin_rows):
    """每行当日总股本（股）：股本变动事件阶梯，当日=最近一次变动生效日≤当日的事件值；
    早于首条事件用首条（上市前无行情行）。share_rows 为空时回退最新已公告财报的
    TOTAL_SHARE（该字段为接口回填的当前股本，仅应急）"""
    import bisect
    share_dates = [s["date"] for s in share_rows] if share_rows else []
    share_vals = [s["total_share"] for s in share_rows] if share_rows else []
    fins = sorted(fin_rows, key=lambda x: (x["notice_date"], x["report_date"]))
    fin_dates = [f["notice_date"] for f in fins]
    out = []
    for r in rows:
        d = r["date"]
        if share_rows:
            i = bisect.bisect_right(share_dates, d) - 1
            out.append(share_vals[i] if i >= 0 else share_vals[0])
            continue
        j = bisect.bisect_right(fin_dates, d) - 1
        out.append(fins[j].get("share") if j >= 0 else None)
    return out

def calc_pe(rows, fin_rows, share_rows):
    """返回 (pe_ttm[], pe_dyn[])。口径同东财行情页：
    PE = 总市值(收盘价 × 当日总股本) ÷ 归母净利润；
    TTM 净利见 calc_ttm_np；动态净利 = 最新累计归母净利 × 季度年化系数"""
    ttm = calc_ttm_np(rows, fin_rows)
    share = calc_share(rows, share_rows, fin_rows)
    import bisect
    fins = sorted(fin_rows, key=lambda x: (x["notice_date"], x["report_date"]))
    dates = [f["notice_date"] for f in fins]
    qmap = {"03-31": 4, "06-30": 2, "09-30": 4.0 / 3, "12-31": 1}
    pe_ttm, pe_dyn = [], []
    for k, r in enumerate(rows):
        d = r["date"]
        close = r.get("close") or 0
        i = bisect.bisect_right(dates, d) - 1
        np_i = fins[i].get("np") if i >= 0 else None
        if i < 0 or not close or not share[k] or np_i is None:
            pe_ttm.append(None); pe_dyn.append(None)
            continue
        latest = fins[i]
        dyn_np = np_i * qmap.get(latest["report_date"][5:10], 1)
        mcap = close * share[k]
        pe_ttm.append(round(mcap / ttm[k], 2) if ttm[k] else None)
        pe_dyn.append(round(mcap / dyn_np, 2) if dyn_np else None)
    return pe_ttm, pe_dyn

# ── PEG：PE-TTM ÷ TTM归母净利同比增速(%) ────────────────────────────
def calc_peg(rows, fin_rows, share_rows, pe_ttm):
    """PEG = PE-TTM ÷ G，G = (当日TTM归母净利 ÷ 一年前TTM归母净利 − 1) × 100
    （净利口径，与市值/净利PE一致；股本变动不影响净利增速）
    G ≤ 0（业绩下滑）或任一缺失 → None（PEG 对负增长无意义）；双指针 O(n)"""
    ttm = calc_ttm_np(rows, fin_rows)
    out = []
    j = 0  # 一年前日期的最大索引（单调不减）
    for i, r in enumerate(rows):
        if pe_ttm[i] is None or ttm[i] is None:
            out.append(None)
            continue
        d = r["date"]
        target = f"{int(d[:4]) - 1:04d}-{d[5:10]}"   # 去年同月同日
        while j + 1 < i and rows[j + 1]["date"] <= target:
            j += 1
        if j >= i or rows[j]["date"] > target or ttm[j] is None or ttm[j] <= 0:
            out.append(None)
            continue
        g = (ttm[i] / ttm[j] - 1) * 100
        out.append(round(pe_ttm[i] / g, 2) if g > 0 else None)
    return out

# ── 历史 PB 计算 ──────────────────────────────────────────────────
def calc_pb(rows, fin_rows):
    """PB = 收盘价 ÷ 每股净资产（最新已公告报告期 BPS，时点值）。BPS 缺失留 None"""
    import bisect
    fins = sorted(fin_rows, key=lambda x: (x["notice_date"], x["report_date"]))
    dates = [f["notice_date"] for f in fins]
    out = []
    for r in rows:
        d = r["date"]
        close = r.get("close") or 0
        i = bisect.bisect_right(dates, d) - 1
        bps = fins[i]["bps"] if i >= 0 else None
        out.append(round(close / bps, 2) if close and bps else None)
    return out

# ── 报告期指标（ROE/ROA等）：公告日滚动取最新值（时点阶梯）──────────
def calc_ratio(rows, fin_rows, key):
    """每个交易日取最新已公告报告期的 key 值（如 roe/roa）；缺失留 None"""
    import bisect
    fins = sorted(fin_rows, key=lambda x: (x["notice_date"], x["report_date"]))
    dates = [f["notice_date"] for f in fins]
    out = []
    for r in rows:
        d = r["date"]
        i = bisect.bisect_right(dates, d) - 1
        v = fins[i].get(key) if i >= 0 else None
        out.append(round(v, 2) if v is not None else None)
    return out

# ── 历史逐日股息率计算（精确365天自然日窗口）──────────────────────────
def calc_dividend_yield(rows, div_rows):
    """每行股息率(%) = 除权日在(当日-365天, 当日]内的每股派息合计 ÷ 当日收盘 × 100
    （精确自然日窗口，同东财口径；避免旧"年月差<=12"把12个月+数天的分红多留一个月造成
    股息率在窗口边缘一天腰斩的悬崖效应）；窗口内无分红记 0.00"""
    from datetime import date
    events = sorted(((date(*map(int, r["ex_date"].split("-"))), r["bonus10"] / 10.0) for r in div_rows))
    out = []
    for r in rows:
        d = date(*map(int, r["date"].split("-")))
        total = 0.0
        for ex, per in events:
            if ex > d:
                continue
            if (d - ex).days < 365:
                total += per
        if not total and events:
            # 空窗回退：窗口内无除权（除权日顺延/披露延迟）时，用最近一次已除权单次派息，
            # 防高股息股在除权空窗期股息率假归零（如粤高速 2026 除权日 8-10，8月初 dy 归 0）
            last = [per for ex, per in events if ex <= d]
            if last:
                total = last[0]
        close = r.get("close") or 0
        out.append(round(total / close * 100, 2) if close and total else 0.0)
    return out

# ── 拉取（不复权）───────────────────────────────────────────────
def fetch_kline(tcode, code, full=True):
    """不复权日K。full=True: 翻页全量+过滤<2004-01-01；full=False: 仅最近800条（增量用）"""
    all_rows = []
    end = ""
    for page in range(30):
        # 最后参数空=不复权；start 参数腾讯忽略，故翻页一律用 end（向前翻）
        param = f"{tcode},day,,{end},800," if end else f"{tcode},day,,,800,"
        url = f"https://web.ifzq.gtimg.cn/appstock/app/fqkline/get?param={param}"
        req = urllib.request.Request(url, headers={"User-Agent": UA})
        d = json.loads(urllib.request.urlopen(req, timeout=15).read().decode("utf-8"))
        k = d.get("data", {}).get(tcode, {})
        days = k.get("day") or k.get("qfqday") or []
        if not days:
            break
        rows = []
        for r in days:
            # 第7位可能是成交额(数字)或分红除权信息(dict)，dict 忽略
            amt = None
            if len(r) > 6 and isinstance(r[6], (str, int, float)):
                try:
                    amt = float(r[6])
                except (TypeError, ValueError):
                    amt = None
            rows.append({"date": r[0], "open": float(r[1]), "close": float(r[2]),
                         "high": float(r[3]), "low": float(r[4]), "volume": float(r[5]),
                         "amount": amt})
        old_first = rows[0]["date"]
        all_rows = rows + all_rows
        if not full:
            break  # 增量：只取最近800条，不翻页
        if len(days) < 800 or old_first == end:
            break
        if old_first < START:
            break  # 已翻过 2004 边界，停止（下方统一过滤）
        end = old_first
        time.sleep(0.5)
    if full:
        all_rows = [r for r in all_rows if r["date"] >= START]
    # 去重排序
    seen = {}
    for r in all_rows:
        seen[r["date"]] = r
    return [seen[k] for k in sorted(seen)]

def make_fetcher(tcode, code):
    def fetcher(last_date):
        if last_date:
            return fetch_kline(tcode, code, full=False)   # 增量：最近800条覆盖合并
        return fetch_kline(tcode, code, full=True)        # 全量：翻页 + 2004过滤
    return fetcher

# ── Excel ───────────────────────────────────────────────────────
def export_excel():
    from datetime import datetime
    from openpyxl import load_workbook
    from _fetch_history import load_cache
    # 展示口径：价格(元)、成交量(万手)、成交额(亿元)；腾讯原始 volume=手、amount=元(估算)
    COL_CN = {"open": "开盘(元)", "close": "收盘(元)", "high": "最高(元)", "low": "最低(元)",
              "volume": "成交量(万手)", "amount": "成交额(亿元)",
              "chg30": "30日涨跌(%)", "chg60": "60日涨跌(%)", "chg90": "90日涨跌(%)"}
    COLS = ["开盘(元)", "收盘(元)", "股息率(%)", "PE(TTM)(倍)", "PE动(倍)", "PB(倍)", "PEG", "ROE(%)", "ROA(%)",
            "最高(元)", "最低(元)", "成交量(万手)", "成交额(亿元)",
            "30日涨跌(%)", "60日涨跌(%)", "90日涨跌(%)"]
    rows_map = {}
    for code, name, _t in STOCKS:
        c = load_cache("股票", code)
        if c:
            fh.fill_etf_amount(c["rows"])          # 估算成交额（元）
            fh.fill_chg_n(c["rows"])               # 30/60/90 交易日涨跌幅（交易日口径）
            fh.save_cache("股票", code, c)          # 写回缓存
            div = load_cache("分红", code)
            fin = load_cache("财报", code)
            share = load_cache("股本", code)
            rows_map[code] = {"name": c.get("name", name), "rows": c["rows"],
                              "div": div["rows"] if div else [],
                              "fin": fin["rows"] if fin else [],
                              "share": share["rows"] if share else []}
    try:
        with pd.ExcelWriter(os.path.join(BASE, "excel", "股票历史.xlsx"), engine="openpyxl") as w:
            for code, info in rows_map.items():
                rows = info["rows"]
                if not rows:
                    continue
                df = pd.DataFrame(rows)
                df["股息率(%)"] = calc_dividend_yield(rows, info["div"])
                pe_ttm, pe_dyn = calc_pe(rows, info["fin"], info["share"])
                df["PE(TTM)(倍)"] = pe_ttm
                df["PE动(倍)"] = pe_dyn
                df["PB(倍)"] = calc_pb(rows, info["fin"])
                df["PEG"] = calc_peg(rows, info["fin"], info["share"], pe_ttm)
                df["ROE(%)"] = calc_ratio(rows, info["fin"], "roe")
                df["ROA(%)"] = calc_ratio(rows, info["fin"], "roa")
                df["date"] = pd.to_datetime(df["date"])
                df = df.sort_values("date").set_index("date")
                df = df.rename(columns=COL_CN)
                df = df[COLS]
                df["成交量(万手)"] = (df["成交量(万手)"] / 1e4).round(2)
                df["成交额(亿元)"] = (df["成交额(亿元)"] / 1e8).round(2)
                df.index.name = "日期"
                df.to_excel(w, sheet_name=f"{code} {info['name'][:10]}"[:31])
        # 日期列显示为年月日 + 美化（列宽/冻结首行/筛选/居中）——导出后独立后处理，避免 pandas 保存覆盖格式
        wb = load_workbook(os.path.join(BASE, "excel", "股票历史.xlsx"))
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
                for cell in row:
                    if isinstance(cell.value, datetime):
                        cell.number_format = "yyyy-mm-dd"
            fh.beautify_sheet(ws)
        wb.save(os.path.join(BASE, "excel", "股票历史.xlsx"))
        print(f"✅ excel/股票历史.xlsx 已生成（{len(rows_map)} 只，2004-01-01 起，不复权，含股息率）")
    except PermissionError:
        print("⚠️  excel/股票历史.xlsx 被占用（可能已在Excel中打开），请关闭后重新执行导出")

def check_financials(codes=None):
    """检测分红/财报是否有更新（如新财报公告/新除权日）：
    拉最新数据与缓存对比（分红比最新除权日，财报比最新公告日），有变化则写回缓存。
    每只约 2 个请求（东财限流 1s）；适合季度末/定期执行。codes=None 默认推荐20只。"""
    lst = codes if codes is not None else [c for c, _, _ in STOCKS]
    names = {c: n for c, n, _ in STOCKS}
    print(f"── 分红/财报更新检测（东财，2s/只，{len(lst)} 只，约 {len(lst)*2//60} 分钟）──")
    updated = []
    for code in lst:
        name = names.get(code, code)
        for typ, key, fetch in (("分红", "ex_date", fetch_dividend), ("财报", "notice_date", fetch_financials)):
            try:
                rows = fetch(code)
                p = fh.cache_path(typ, code)
                changed = not os.path.exists(p)
                if not changed:
                    old = json.load(open(p, encoding="utf-8"))
                    old_rows = old.get("rows", [])
                    new_v = rows[0].get(key) if rows else None
                    old_v = old_rows[0].get(key) if old_rows else None
                    changed = new_v != old_v or len(rows) != len(old_rows)
                if changed:
                    fh.save_cache(typ, code, {"code": code, "name": name,
                                              "fetched_at": time.strftime("%Y-%m-%d"), "rows": rows})
                    updated.append(f"{typ} {code} {name}（{len(rows)}条）")
            except Exception as e:
                print(f"  ❌ [{code} {name}] {typ}检测失败: {repr(e)[:60]}")
            time.sleep(0.3)
    if updated:
        print(f"  📦 更新 {len(updated)} 项：{'；'.join(updated)}")
    else:
        print("  ✅ 分红/财报均无更新")
    return updated

# ── 主流程 ──────────────────────────────────────────────────────
def update_all(refresh=False, refresh_fin=False):
    print("═══ 推荐20只股票历史行情（不复权，2004-01-01起）═══")
    if refresh_fin:
        check_financials()
        return
    update_dividends(refresh=refresh)   # 分红缓存缺失才拉（秒级），删除自愈；refresh 强制重拉
    update_financials(refresh=refresh)  # 财报缓存缺失才拉，删除自愈
    for code, name, tcode in STOCKS:
        if refresh:
            rows = fetch_kline(tcode, code, full=True)
            obj = {"code": code, "name": name, "fetched_at": time.strftime("%Y-%m-%d"), "rows": rows}
            fh.save_cache("股票", code, obj)
            print(f"  [{code} {name}] 全量刷新 {len(rows)}条 -> cache/股票_{code}.json")
        else:
            try:
                fh.update_incremental("股票", code, name, make_fetcher(tcode, code))
            except Exception as e:
                print(f"  ❌ [{code} {name}] 失败: {repr(e)[:80]}")
        time.sleep(0.3)
    export_excel()
    print("\n✅ 股票历史更新完成")

if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--refresh", action="store_true", help="忽略缓存，全量重拉")
    ap.add_argument("--refresh-fin", action="store_true", help="检测分红/财报更新（对比最新除权日/公告日，有变化才写缓存）")
    args = ap.parse_args()
    update_all(refresh=args.refresh, refresh_fin=args.refresh_fin)
