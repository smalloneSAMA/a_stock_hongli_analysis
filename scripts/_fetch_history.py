# -*- coding: utf-8 -*-
"""
红利指数与ETF历史行情拉取工具
- 指数历史日线：腾讯K线(000922/000015/000821/000825/000151) + 中证官网index-perf(H30269/930955/932315/931468/H30270) + 国证官网(980092)
- ETF历史：腾讯K线(场内价格) + 新浪净值(单位/累计净值)
- 缓存：cache/{类型}_{代码}.json，有缓存直接复用；--refresh 强制重新拉取
- 输出：excel/指数历史.xlsx、excel/ETF历史.xlsx（每标的一个sheet）
"""
import sys, io, os, json, time, argparse, urllib.request, requests, datetime
import pandas as pd

if __name__ == "__main__":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # 项目根（scripts/的父目录）
CACHE_DIR = os.path.join(BASE, "cache")
EXCEL_DIR = os.path.join(BASE, "excel")
os.makedirs(CACHE_DIR, exist_ok=True)
os.makedirs(EXCEL_DIR, exist_ok=True)

UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"
S = requests.Session()
S.headers.update({"User-Agent": UA})

# ── 标的清单 ───────────────────────────────────────────────────────
INDICES = [
    ("000922", "中证红利", "tencent", "sh000922"),
    ("000015", "上证红利", "tencent", "sh000015"),
    ("000821", "沪深300红利", "tencent", "sh000821"),
    ("000825", "中证央企红利", "tencent", "sh000825"),
    ("H30269", "中证红利低波动", "csindex", None),
    ("930955", "红利低波动100", "csindex", None),
    ("932315", "中证全指红利质量", "csindex", None),
    ("931468", "中证红利质量", "csindex", None),
    ("H30270", "中证红利价值", "csindex", None),
    ("980092", "国证自由现金流", "cnindex", None),
    ("000151", "上证国有企业红利(辅助,510720跟踪)", "tencent", "sh000151"),
]
ETFS = [
    ("512890", "红利低波ETF华泰柏瑞", "sh512890"),
    ("515180", "红利ETF易方达", "sh515180"),
    ("563020", "红利低波ETF易方达", "sh563020"),
    ("515080", "中证红利ETF招商", "sh515080"),
    ("159549", "红利低波ETF天弘", "sz159549"),
    ("561580", "央企红利ETF华泰柏瑞", "sh561580"),
    ("510720", "红利国企ETF国泰", "sh510720"),
    ("159209", "红利质量ETF招商", "sz159209"),
    ("159758", "华夏中证红利质量ETF", "sz159758"),
    ("563700", "红利价值ETF易方达", "sh563700"),
    ("159229", "自由现金流ETF广发", "sz159229"),
]

# ── 缓存 ───────────────────────────────────────────────────────────
def cache_path(typ, code):
    return os.path.join(CACHE_DIR, f"{typ}_{code}.json")

def load_cache(typ, code):
    p = cache_path(typ, code)
    if os.path.exists(p):
        with open(p, encoding="utf-8") as f:
            return json.load(f)
    return None

def save_cache(typ, code, obj):
    p = cache_path(typ, code)
    tmp = p + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False)
    os.replace(tmp, p)  # 原子写：先写临时文件再替换，防中断损坏缓存

# ── 1. 腾讯K线（全历史或增量，翻页）────────────────────────────
def fetch_tencent_kline(tcode, code, start=None):
    """tcode: sh000922; start=起始日期(YYYY-MM-DD,含), 默认全量; 返回 [{date, open, close, high, low, volume, amount}]"""
    all_rows = []
    end = ""
    for page in range(30):
        if start and not end:
            param = f"{tcode},day,{start},,800,qfq"
        else:
            param = f"{tcode},day,,{end},800,qfq"
        url = f"https://web.ifzq.gtimg.cn/appstock/app/fqkline/get?param={param}"
        req = urllib.request.Request(url, headers={"User-Agent": UA})
        d = json.loads(urllib.request.urlopen(req, timeout=15).read().decode("utf-8"))
        k = d.get("data", {}).get(tcode, {})
        days = k.get("day") or k.get("qfqday") or []
        if not days:
            break
        # 腾讯日K: [date, open, close, high, low, volume, ...]（可能带amount在第7位）
        rows = []
        for r in days:
            amt = float(r[6]) if len(r) > 6 and r[6] else None
            rows.append({"date": r[0], "open": float(r[1]), "close": float(r[2]),
                         "high": float(r[3]), "low": float(r[4]), "volume": float(r[5]),
                         "amount": amt})
        old_first = rows[0]["date"]
        all_rows = rows + all_rows
        if len(days) < 800 or old_first == end:
            break
        end = old_first
        time.sleep(0.5)
    return all_rows

# ── 2. 中证官网 index-perf（H30269等5只）─────────────────────────
def fetch_csindex_perf(code, start=None, end=None):
    """start/end: YYYYMMDD，None=最早/最新；返回 [{tradeDate, open, high, low, close, change, changePct, tradingVol, tradingValue}]"""
    url = "https://www.csindex.com.cn/csindex-home/perf/index-perf"
    all_rows = []
    # 按年分段拉取，避免接口单次数量上限；end 缺省取今天（接口返回截至当日数据）
    end_year = int((end or datetime.date.today().strftime("%Y%m%d"))[:4])
    start_year = int((start or "20040101")[:4])
    for y in range(start_year, end_year + 1):
        s = start if y == start_year else f"{y}0101"
        e = end if y == end_year else f"{y}1231"
        rows = []
        for attempt in range(4):
            params = {"indexCode": code, "startDate": s, "endDate": e}
            try:
                r = S.get(url, params=params, timeout=15)
                d = r.json()
                rows = d.get("data") or []
                break
            except Exception:
                rows = []
                time.sleep(2 + attempt * 2)
        if rows:
            all_rows.extend(rows)
        time.sleep(1.5 + (y % 3) * 0.5)  # 间隔拉长防限流
    out = []
    for r in all_rows:
        out.append({"date": r["tradeDate"], "open": r.get("open"), "high": r.get("high"),
                    "low": r.get("low"), "close": r.get("close"), "volume": r.get("tradingVol"),
                    "amount": r.get("tradingValue")})
    # 按日期去重排序
    seen = {}
    for r in out:
        d = r["date"]
        if len(d) == 8:
            d = f"{d[:4]}-{d[4:6]}-{d[6:]}"
            r["date"] = d
        seen[d] = r
    return [seen[k] for k in sorted(seen)]

# ── 3. 国证官网 K线（980092）─────────────────────────────────────
def fetch_cnindex_kline(code, start=None):
    """start=YYYY-MM-DD，默认全量；返回 [{date, current, high, open, low, close, chg, percent, amount, volume}]"""
    url = "https://hq.cnindex.com.cn/market/market/getIndexDailyDataWithDataFormat"
    all_rows = []
    start_year = int((start or "2012-01-01")[:4])
    for y in range(start_year, datetime.date.today().year + 1):
        s = start if y == start_year else f"{y}-01-01"
        params = {"indexCode": code, "startDate": s, "endDate": f"{y}-12-31", "frequency": "day"}
        try:
            r = S.get(url, params=params, timeout=20)
            d = r.json()
            item = d.get("data", {}).get("item", [])
            data = d.get("data", {}).get("data", [])
        except Exception:
            data = []
        for row in data:
            rec = dict(zip(item, row))
            all_rows.append({"date": rec["timestamp"][:10], "open": rec.get("open"),
                             "high": rec.get("high"), "low": rec.get("low"),
                             "close": rec.get("close"), "volume": rec.get("volume"),
                             "amount": rec.get("amount")})
        time.sleep(0.6)
    seen = {}
    for r in all_rows:
        seen[r["date"]] = r
    return [seen[k] for k in sorted(seen)]

# ── 4. 新浪净值 ───────────────────────────────────────────────────
def fetch_sina_nav(code, start=None):
    """start=YYYY-MM-DD(含)；返回 [{date, nav(单位净值), acc_nav(累计净值)}]
    ⚠️ 新浪 datefrom 参数实测不生效（返回全量倒序），本地过滤 start 之后的数据"""
    all_rows = []
    page = 1
    while page <= 50:
        url = "https://stock.finance.sina.com.cn/fundInfo/api/openapi.php/CaihuiFundInfoService.getNav"
        r = S.get(url, params={"symbol": code, "datefrom": start or "", "dateto": "", "page": page, "num": 1000}, timeout=15)
        d = r.json()
        rows = d.get("result", {}).get("data", {}).get("data") or []
        if not rows:
            break
        for x in rows:
            dt = x["fbrq"][:10]
            if start and dt < start:
                continue  # 本地过滤（新浪datefrom不生效）
            all_rows.append({"date": dt, "nav": float(x["jjjz"]), "acc_nav": float(x["ljjz"])})
        if len(rows) < 1000:
            break
        page += 1
        time.sleep(0.6)
    seen = {}
    for r in all_rows:
        seen[r["date"]] = r
    return [seen[k] for k in sorted(seen)]

# ── 增量更新 ────────────────────────────────────────────────────
def update_incremental(typ, code, name, fetcher):
    """增量更新：读缓存最后日期，只拉之后的数据并合并；无缓存则全量。返回 (新增条数, 总条数, 最后日期)"""
    cached = load_cache(typ, code)
    old_rows = {r["date"]: r for r in cached["rows"]} if cached else {}
    last_date = max(old_rows) if old_rows else None
    if last_date:
        print(f"  [{code} {name}] 增量更新（从 {last_date} 起）...")
    else:
        print(f"  [{code} {name}] 无缓存，全量拉取...")
    rows = fetcher(last_date)
    n_new = 0
    for r in rows:
        d = r["date"]
        if not d:
            continue
        if d in old_rows:
            # 字段级合并：新值覆盖旧值，但新行为空(None)的字段保留旧值，
            # 防止“增量返回不完整行”整体覆盖掉缓存里的好数据
            merged = dict(old_rows[d])
            for k, v in r.items():
                if k != "date" and v is not None:
                    merged[k] = v
            if merged != old_rows[d]:
                n_new += 1
            old_rows[d] = merged
        else:
            old_rows[d] = r
            n_new += 1
    merged = [old_rows[k] for k in sorted(old_rows)]
    if typ == "指数":
        merged = [{k: r.get(k) for k in ("date", "open", "close", "high", "low", "volume", "amount")} for r in merged]
    obj = {"code": code, "name": name, "fetched_at": time.strftime("%Y-%m-%d"), "rows": merged}
    save_cache(typ, code, obj)
    print(f"  [{code} {name}] 新增 {n_new}条，累计 {len(merged)}条，最新 {merged[-1]['date']}")
    return n_new, len(merged), merged[-1]["date"]

# ── ETF成交额估算 ───────────────────────────────────────────────
def fill_etf_amount(rows):
    """腾讯K线无成交额字段：成交额(元) ≈ 成交量(手)×100×(最高+最低+收盘)/3
    实测512890估算值 vs 实际成交额误差<1%"""
    n = 0
    for r in rows:
        if r.get("amount") is None and r.get("volume") and r.get("close"):
            h = r.get("high") or r["close"]
            l = r.get("low") or r["close"]
            c = r["close"]
            r["amount"] = round(r["volume"] * 100 * (h + l + c) / 3, 2)
            n += 1
    return rows

# ── 主流程 ────────────────────────────────────────────────────────
def get_or_fetch(typ, code, name, fetcher, refresh):
    if not refresh:
        cached = load_cache(typ, code)
        if cached:
            print(f"  [{code} {name}] 使用缓存（{cached['fetched_at']}，{len(cached['rows'])}条）")
            return cached
    print(f"  [{code} {name}] 拉取中...")
    rows = fetcher()
    obj = {"code": code, "name": name, "fetched_at": time.strftime("%Y-%m-%d"),
           "rows": rows}
    save_cache(typ, code, obj)
    print(f"  [{code} {name}] 完成 {len(rows)}条 -> cache/{typ}_{code}.json")
    return obj

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--refresh", action="store_true", help="忽略缓存，强制重新拉取")
    ap.add_argument("--type", choices=["index", "etf", "all"], default="all")
    args = ap.parse_args()

    idx_data, etf_data = {}, {}
    if args.type in ("index", "all"):
        print("═══ 指数历史（日线）═══")
        for code, name, src, tcode in INDICES:
            def fetcher(src=src, code=code, tcode=tcode):
                if src == "tencent":
                    return fetch_tencent_kline(tcode, code)
                if src == "csindex":
                    return fetch_csindex_perf(code)
                return fetch_cnindex_kline(code)
            obj = get_or_fetch("指数", code, name, fetcher, args.refresh)
            idx_data[code] = {"name": name, "source": src, "rows": obj["rows"]}

    if args.type in ("etf", "all"):
        print("═══ ETF历史（场内价格+净值）═══")
        for code, name, tcode in ETFS:
            def fetcher(code=code, tcode=tcode):
                kline = fetch_tencent_kline(tcode, code)
                nav = fetch_sina_nav(code)
                # merge：kline主表 + 净值
                kmap = {r["date"]: r for r in kline}
                nmap = {r["date"]: r for r in nav}
                rows = []
                for d in sorted(set(kmap) | set(nmap)):
                    r = {"date": d}
                    if d in kmap:
                        r.update({k: v for k, v in kmap[d].items() if k != "date"})
                    if d in nmap:
                        r["nav"] = nmap[d]["nav"]
                        r["acc_nav"] = nmap[d]["acc_nav"]
                    rows.append(r)
                return rows
            obj = get_or_fetch("ETF", code, name, fetcher, args.refresh)
            etf_data[code] = {"name": name, "rows": obj["rows"]}

    # ── 写Excel ──────────────────────────────────────────────────
    COL_CN = {
        "open": "开盘", "close": "收盘", "high": "最高", "low": "最低",
        "volume": "成交量", "amount": "成交额", "change": "涨跌额", "changePct": "涨跌幅",
        "nav": "单位净值", "acc_nav": "累计净值",
    }
    if idx_data:
        with pd.ExcelWriter(os.path.join(EXCEL_DIR, "指数历史.xlsx"), engine="openpyxl") as w:
            for code, info in idx_data.items():
                rows = info["rows"]
                if not rows:
                    continue
                df = pd.DataFrame(rows)
                df["date"] = pd.to_datetime(df["date"])
                df = df.sort_values("date").set_index("date")
                df = df.rename(columns=COL_CN)
                df = df[["开盘", "收盘", "最高", "最低", "成交量", "成交额"]]
                df.index.name = "日期"
                src_cn = {"tencent": "腾讯K线", "csindex": "中证官网", "cnindex": "国证官网"}[info["source"]]
                sheet = f"{code} {info['name'][:10]}"
                df.to_excel(w, sheet_name=sheet[:31])
            print(f"✅ excel/指数历史.xlsx（数据源：腾讯K线/中证官网/国证官网）")

    if etf_data:
        with pd.ExcelWriter(os.path.join(EXCEL_DIR, "ETF历史.xlsx"), engine="openpyxl") as w:
            for code, info in etf_data.items():
                rows = fill_etf_amount(info["rows"])
                if not rows:
                    continue
                df = pd.DataFrame(rows)
                df["date"] = pd.to_datetime(df["date"])
                df = df.sort_values("date").set_index("date")
                df = df.rename(columns=COL_CN)
                if "单位净值" in df.columns:
                    df = df[["开盘", "收盘", "最高", "最低", "成交量", "成交额", "单位净值", "累计净值"]]
                df.index.name = "日期"
                sheet = f"{code} {info['name'][:10]}"
                df.to_excel(w, sheet_name=sheet[:31])
            print("✅ excel/ETF历史.xlsx（场内价格=腾讯K线，净值=新浪）")

    print("全部完成。缓存目录: cache/ · Excel目录: excel/")

# ── Excel 美化（列宽自适应/冻结首行/筛选/居中）────────────────────
def beautify_sheet(ws):
    """单个sheet：列宽按内容显示宽度自适应、冻结首行、开启筛选、全表水平垂直居中，表头加粗浅蓝底。
    幂等，可重复执行"""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    if ws.max_row < 1 or ws.max_column < 1:
        return
    nrow, ncol = ws.max_row, ws.max_column

    def dispw(v):
        if isinstance(v, datetime.datetime):
            return 10   # 日期按 yyyy-mm-dd 显示宽度计
        s = str(v)
        return sum(2 if ord(ch) > 127 else 1 for ch in s)   # 中文/全角按2字符宽

    widths = [0] * ncol
    for row in ws.iter_rows(min_row=1, max_row=nrow, max_col=ncol):
        for cell in row:
            if cell.value is None:
                continue
            w = dispw(cell.value)
            i = cell.column - 1
            if w > widths[i]:
                widths[i] = w
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 8), 22)

    ws.freeze_panes = "A2"            # 冻结首行
    ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}{nrow}"   # 开启筛选

    center = Alignment(horizontal="center", vertical="center")
    hfont = Font(bold=True)
    hfill = PatternFill("solid", fgColor="D9E1F2")
    for row in ws.iter_rows(min_row=1, max_row=nrow, max_col=ncol):
        for cell in row:
            if cell.value is None:
                continue
            cell.alignment = center
            if cell.row == 1:
                cell.font = hfont
                cell.fill = hfill

def beautify_file(path):
    """打开xlsx逐sheet美化后保存（幂等）"""
    from openpyxl import load_workbook
    wb = load_workbook(path)
    for ws in wb.worksheets:
        beautify_sheet(ws)
    wb.save(path)

if __name__ == "__main__":
    main()
